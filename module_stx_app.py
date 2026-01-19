import openpyxl

import homepage_driver
# import login_app
import var_stx_app
import re
import caseid_stx_app
import module_other_stx_app












def ModuleTest():
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))
    for i in modetest:
        print("so1", i)
        if i == "0":
            luong_all(self='')
        if i == "1":
            luong_customer(self='')
        if i == "2":
            luong_drive(self='')





#case tương tác 2 app
def get_cases_interaction():
    list_caseinteraction = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "1"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["H"+rownum].value == "x"):
                    case_fail = sheet["A"+rownum].value
                    list_caseinteraction.append(case_fail)
                rownum = int(rownum)

    count = len(list_caseinteraction)
    print(f"Case tương tác 2 app: {list_caseinteraction}")
    print(f"Tổng số case tương tác 2 app: {count}")
    return list_caseinteraction



#case customer
def get_cases_customer():
    list_casecustomer = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "2"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["I"+rownum].value == "x"):
                    case_fail = sheet["A"+rownum].value
                    list_casecustomer.append(case_fail)
                rownum = int(rownum)

    count = len(list_casecustomer)
    print(f"Case khách hàng: {list_casecustomer}")
    print(f"Tổng số case khách hàng: {count}")
    return list_casecustomer


#case_driver
def get_cases_driver():
    list_casedriver = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "3"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["J"+rownum].value == "x"):
                    case_fail = sheet["A"+rownum].value
                    list_casedriver.append(case_fail)
                rownum = int(rownum)

    count = len(list_casedriver)
    print(f"Case lái xe: {list_casedriver}")
    print(f"Tổng số case lái xe: {count}")
    return list_casedriver





#case fail tương tác 2 app
def get_cases_fail_interaction():
    list_caseinteraction = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "1"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["H"+rownum].value == "x") and (sheet["G"+rownum].value == "Fail"):
                    case_fail = sheet["A"+rownum].value
                    list_caseinteraction.append(case_fail)
                rownum = int(rownum)

    count = len(list_caseinteraction)
    print(f"Case fail tương tác 2 app: {list_caseinteraction}")
    print(f"Tổng số fail case tương tác 2 app: {count}")
    return list_caseinteraction


#case fail customer
def get_cases_fail_customer():
    list_casecustomer = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "2"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["I"+rownum].value == "x") and (sheet["G"+rownum].value == "Fail"):
                    case_fail = sheet["A"+rownum].value
                    list_casecustomer.append(case_fail)
                rownum = int(rownum)

    count = len(list_casecustomer)
    print(f"Case fail khách hàng: {list_casecustomer}")
    print(f"Tổng số case fail khách hàng: {count}")
    return list_casecustomer


#case_fail driver
def get_cases_fail_driver():
    list_casedriver = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "3"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["J"+rownum].value == "x") and (sheet["G"+rownum].value == "Fail"):
                    case_fail = sheet["A"+rownum].value
                    list_casedriver.append(case_fail)
                rownum = int(rownum)

    count = len(list_casedriver)
    print(f"Case fail lái xe: {list_casedriver}")
    print(f"Tổng số case fail lái xe: {count}")
    return list_casedriver









lis_casenone_skip = ['Account12_1', 'Account12_2', 'Account12_3','Account12_4', 'Account12_5',
                     'Account12_6', 'Account18', 'Home24', 'History15', 'History16', 'History17']

#case none tương tác 2 app
def get_cases_none_interaction():
    list_caseinteraction = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "1"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["H"+rownum].value == "x") and (sheet["G"+rownum].value == None):
                    case_fail = sheet["A"+rownum].value
                    if case_fail not in lis_casenone_skip:
                        list_caseinteraction.append(case_fail)
                rownum = int(rownum)

    count = len(list_caseinteraction)
    print(f"Case none tương tác 2 app: {list_caseinteraction}")
    print(f"Tổng số none case tương tác 2 app: {count}")
    return list_caseinteraction



#case none customer
def get_cases_none_customer():
    list_casecustomer = []


    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "2"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["I"+rownum].value == "x") and (sheet["G"+rownum].value == None):
                    case_fail = sheet["A"+rownum].value
                    if case_fail not in lis_casenone_skip:
                        list_casecustomer.append(case_fail)
                rownum = int(rownum)

    count = len(list_casecustomer)
    print(f"Case None khách hàng: {list_casecustomer}")
    print(f"Tổng số case None khách hàng: {count}")
    return list_casecustomer



#case none driver
def get_cases_none_driver():
    list_casedriver = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if (i == "0") or (i == "3"):
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if (sheet["J"+rownum].value == "x") and (sheet["G"+rownum].value == None):
                    case_fail = sheet["A"+rownum].value
                    if case_fail not in lis_casenone_skip:
                        list_casedriver.append(case_fail)
                rownum = int(rownum)

    count = len(list_casedriver)
    print(f"Case None lái xe: {list_casedriver}")
    print(f"Tổng số case None lái xe: {count}")
    return list_casedriver














def check_casenone():
    module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 86, 2, 0)

    list_casenone = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)

                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)

                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)

                rownum = int(rownum)
            print(list_casenone)
            count = len(list_casenone)
            print("Số case trống luồng full(0): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 86, 2, count)


        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)
                rownum = int(rownum)
            print(list_casenone)
            count = len(list_casenone)
            print("Số case trống Luồng tương tác 2app(1): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 86, 2, count)



        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)
                rownum = int(rownum)
            print(list_casenone)
            count = len(list_casenone)
            print("Số case trống Luồng khách hàng(2): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 86, 2, count)



        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casenone.append(case_fail)
                rownum = int(rownum)
            print(list_casenone)
            count = len(list_casenone)
            print("Số case trống Luồng tài xế(3): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 86, 2, count)


    return list_casenone




def check_casefail():
    module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 167, 2, 0)
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if i == "0":
            while (rownum < 600):
                rownum += 1
                rownum = str(rownum)
                # print(sheet["G"+rownum].value)
                # print(sheet["H"+rownum].value)

                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)

                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)

                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)


                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case fail luồng full(0): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 167, 2, count)


        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case fail Luồng tương tác 2app(1): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 167, 2, count)



        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case fail Luồng khách hàng(2): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 167, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case fail Luồng tài xế(3): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 167, 2, count)


    return list_casefail




def check_casepass():
    module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 177, 2, 0)
    list_casepass = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))

    for i in modetest:
        print("i", i)
        if i == "0":
            while (rownum < 600):
                rownum += 1
                rownum = str(rownum)
                # print(sheet["G"+rownum].value)
                # print(sheet["H"+rownum].value)

                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)

                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)

                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    # print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)


                rownum = int(rownum)
            print(list_casepass)
            count = len(list_casepass)
            print("Số case pass luồng full(0): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 177, 2, count)


        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)
                rownum = int(rownum)
            print(list_casepass)
            count = len(list_casepass)
            print("Số case pass Luồng tương tác 2app(1): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 177, 2, count)



        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)
                rownum = int(rownum)
            print(list_casepass)
            count = len(list_casepass)
            print("Số case pass Luồng khách hàng(2): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 177, 2, count)



        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casepass.append(case_fail)
                rownum = int(rownum)
            print(list_casepass)
            count = len(list_casepass)
            print("Số case pass Luồng tài xế(3): ", count)
            module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 177, 2, count)

    return list_casepass





def check_casespecial():
    module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 85, 2, 0)
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")

    rownum = 9
    # Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    module_other_stx_app.writeData1(var_stx_app.path_luutamthoi, "Sheet1", 85, 2, count)
















def retest_casefail_none(self):
    import homepage_g7
    import homepage_driver


    # module_other_stx_app.reset_driver_app("com.binhanh.g7taxi")
    # homepage_g7.homepage_g7_back()
    #
    # module_other_stx_app.reset_driver_app("com.binhanh.driver.g7")
    # homepage_driver.homepage_g7_drive_back()



    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        if modetest == "0":
            if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                print(sheet["A"+rownum].value)
                case_fail = sheet["A"+rownum].value
                list_casefail.append(case_fail)

        if modetest == "1":
            if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                print(sheet["A"+rownum].value)
                case_fail = sheet["A"+rownum].value
                list_casefail.append(case_fail)

        if modetest == "2":
            if sheet["J" + rownum].value == "x" and sheet["G" + rownum].value == None:
                print(sheet["A" + rownum].value)
                case_fail = sheet["A" + rownum].value
                list_casefail.append(case_fail)

        if modetest == "3":
            if sheet["K" + rownum].value == "x" and sheet["G" + rownum].value == None:
                print(sheet["A" + rownum].value)
                case_fail = sheet["A" + rownum].value
                list_casefail.append(case_fail)

        rownum = int(rownum)
    print(list_casefail)
    for case in list_casefail:
        try:
            if case == 'Login01':
                caseid_stx_app.caseid_login01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login02':
                caseid_stx_app.caseid_login02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login03':
                caseid_stx_app.caseid_login03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login04':
                caseid_stx_app.caseid_login04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login05':
                caseid_stx_app.caseid_login05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login06':
                caseid_stx_app.caseid_login06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login07':
                caseid_stx_app.caseid_login07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login08':
                caseid_stx_app.caseid_login08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login09':
                caseid_stx_app.caseid_login09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login10':
                caseid_stx_app.caseid_login10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'HomePage01':
                caseid_stx_app.caseid_homepage01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage02':
                caseid_stx_app.caseid_homepage02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage03':
                caseid_stx_app.caseid_homepage03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage04':
                caseid_stx_app.caseid_homepage04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage05':
                caseid_stx_app.caseid_homepage05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage06':
                caseid_stx_app.caseid_homepage06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage07':
                caseid_stx_app.caseid_homepage07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage08':
                caseid_stx_app.caseid_homepage08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage09':
                caseid_stx_app.caseid_homepage09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage10':
                caseid_stx_app.caseid_homepage10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage11':
                caseid_stx_app.caseid_homepage11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12':
                caseid_stx_app.caseid_homepage12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage12_1':
                caseid_stx_app.caseid_homepage12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12_2':
                caseid_stx_app.caseid_homepage12_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage13':
                caseid_stx_app.caseid_homepage13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage14':
                caseid_stx_app.caseid_homepage14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage15':
                caseid_stx_app.caseid_homepage15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage16':
                caseid_stx_app.caseid_homepage16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage17':
                caseid_stx_app.caseid_homepage17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage18':
                caseid_stx_app.caseid_homepage18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage19':
                caseid_stx_app.caseid_homepage19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage20':
                caseid_stx_app.caseid_homepage20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage21':
                caseid_stx_app.caseid_homepage21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage22':
                caseid_stx_app.caseid_homepage22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage23':
                caseid_stx_app.caseid_homepage23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage24':
                caseid_stx_app.caseid_homepage24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage25':
                caseid_stx_app.caseid_homepage25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage26':
                caseid_stx_app.caseid_homepage26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage27':
                caseid_stx_app.caseid_homepage27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28':
                caseid_stx_app.caseid_homepage28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28_1':
                caseid_stx_app.caseid_homepage28_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage29':
                caseid_stx_app.caseid_homepage29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage30':
                caseid_stx_app.caseid_homepage30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage31':
                caseid_stx_app.caseid_homepage31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage32':
                caseid_stx_app.caseid_homepage32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage33':
                caseid_stx_app.caseid_homepage33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage34':
                caseid_stx_app.caseid_homepage34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage35':
                caseid_stx_app.caseid_homepage35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage36':
                caseid_stx_app.caseid_homepage36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage37':
                caseid_stx_app.caseid_homepage37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage38':
                caseid_stx_app.caseid_homepage38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage39':
                caseid_stx_app.caseid_homepage39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage40':
                caseid_stx_app.caseid_homepage40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage41':
                caseid_stx_app.caseid_homepage41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42':
                caseid_stx_app.caseid_homepage42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42_1':
                caseid_stx_app.caseid_homepage42_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage43':
                caseid_stx_app.caseid_homepage43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_1':
                caseid_stx_app.caseid_homepage43_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_2':
                caseid_stx_app.caseid_homepage43_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage44':
                caseid_stx_app.caseid_homepage44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage45':
                caseid_stx_app.caseid_homepage45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage46':
                caseid_stx_app.caseid_homepage46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage47':
                caseid_stx_app.caseid_homepage47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage48':
                caseid_stx_app.caseid_homepage48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage49':
                caseid_stx_app.caseid_homepage49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage50':
                caseid_stx_app.caseid_homepage50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage51':
                caseid_stx_app.caseid_homepage51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage52':
                caseid_stx_app.caseid_homepage52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage53':
                caseid_stx_app.caseid_homepage53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage54':
                caseid_stx_app.caseid_homepage54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage55':
                caseid_stx_app.caseid_homepage55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage56':
                caseid_stx_app.caseid_homepage56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage57':
                caseid_stx_app.caseid_homepage57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage58':
                caseid_stx_app.caseid_homepage58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage59':
                caseid_stx_app.caseid_homepage59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage60':
                caseid_stx_app.caseid_homepage60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage61':
                caseid_stx_app.caseid_homepage61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage62':
                caseid_stx_app.caseid_homepage62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage63':
                caseid_stx_app.caseid_homepage63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage64':
                caseid_stx_app.caseid_homepage64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage65':
                caseid_stx_app.caseid_homepage65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage66':
                caseid_stx_app.caseid_homepage66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage67':
                caseid_stx_app.caseid_homepage67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage68':
                caseid_stx_app.caseid_homepage68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage69':
                caseid_stx_app.caseid_homepage69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage70':
                caseid_stx_app.caseid_homepage70(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'FavoriteSpot01':
                caseid_stx_app.caseid_favoritespot01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot02':
                caseid_stx_app.caseid_favoritespot02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot03':
                caseid_stx_app.caseid_favoritespot03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot04':
                caseid_stx_app.caseid_favoritespot04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot05':
                caseid_stx_app.caseid_favoritespot05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot06':
                caseid_stx_app.caseid_favoritespot06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot07':
                caseid_stx_app.caseid_favoritespot07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot08':
                caseid_stx_app.caseid_favoritespot08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot09':
                caseid_stx_app.caseid_favoritespot09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot10':
                caseid_stx_app.caseid_favoritespot10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot11':
                caseid_stx_app.caseid_favoritespot11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot12':
                caseid_stx_app.caseid_favoritespot12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Promotion01':
                caseid_stx_app.caseid_promotion01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion02':
                caseid_stx_app.caseid_promotion02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion03':
                caseid_stx_app.caseid_promotion03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion04':
                caseid_stx_app.caseid_promotion04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion05':
                caseid_stx_app.caseid_promotion05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion06':
                caseid_stx_app.caseid_promotion06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion07':
                caseid_stx_app.caseid_promotion07(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account01':
                caseid_stx_app.caseid_account01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account02':
                caseid_stx_app.caseid_account02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account03':
                caseid_stx_app.caseid_account03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account04':
                caseid_stx_app.caseid_account04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account05':
                caseid_stx_app.caseid_account05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account06':
                caseid_stx_app.caseid_account06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account07':
                caseid_stx_app.caseid_account07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account08':
                caseid_stx_app.caseid_account08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account09':
                caseid_stx_app.caseid_account09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account10':
                caseid_stx_app.caseid_account10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account11':
                caseid_stx_app.caseid_account11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12':
                caseid_stx_app.caseid_account12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_1':
                caseid_stx_app.caseid_account12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_2':
                caseid_stx_app.caseid_account12_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_3':
                caseid_stx_app.caseid_account12_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_4':
                caseid_stx_app.caseid_account12_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_5':
                caseid_stx_app.caseid_account12_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_6':
                caseid_stx_app.caseid_account12_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13':
                caseid_stx_app.caseid_account13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13_1':
                caseid_stx_app.caseid_account13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account14':
                caseid_stx_app.caseid_account14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account15':
                caseid_stx_app.caseid_account15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account16':
                caseid_stx_app.caseid_account16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account17':
                caseid_stx_app.caseid_account17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account18':
                caseid_stx_app.caseid_account18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account19':
                caseid_stx_app.caseid_account19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account20':
                caseid_stx_app.caseid_account20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account21':
                caseid_stx_app.caseid_account21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account22':
                caseid_stx_app.caseid_account22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account23':
                caseid_stx_app.caseid_account23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account24':
                caseid_stx_app.caseid_account24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account25':
                caseid_stx_app.caseid_account25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account26':
                caseid_stx_app.caseid_account26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account27':
                caseid_stx_app.caseid_account27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account28':
                caseid_stx_app.caseid_account28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account29':
                caseid_stx_app.caseid_account29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account30':
                caseid_stx_app.caseid_account30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account31':
                caseid_stx_app.caseid_account31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account32':
                caseid_stx_app.caseid_account32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account33':
                caseid_stx_app.caseid_account33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account34':
                caseid_stx_app.caseid_account34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account35':
                caseid_stx_app.caseid_account35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account36':
                caseid_stx_app.caseid_account36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account37':
                caseid_stx_app.caseid_account37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account38':
                caseid_stx_app.caseid_account38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account39':
                caseid_stx_app.caseid_account39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40':
                caseid_stx_app.caseid_account40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40_1':
                caseid_stx_app.caseid_account40_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account41':
                caseid_stx_app.caseid_account41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account42':
                caseid_stx_app.caseid_account42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account43':
                caseid_stx_app.caseid_account43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account44':
                caseid_stx_app.caseid_account44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account45':
                caseid_stx_app.caseid_account45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account46':
                caseid_stx_app.caseid_account46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account47':
                caseid_stx_app.caseid_account47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account48':
                caseid_stx_app.caseid_account48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account49':
                caseid_stx_app.caseid_account49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50':
                caseid_stx_app.caseid_account50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50_1':
                caseid_stx_app.caseid_account50_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account51':
                caseid_stx_app.caseid_account51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_1':
                caseid_stx_app.caseid_account51_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_2':
                caseid_stx_app.caseid_account51_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account52':
                caseid_stx_app.caseid_account52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account53':
                caseid_stx_app.caseid_account53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account54':
                caseid_stx_app.caseid_account54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account55':
                caseid_stx_app.caseid_account55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account56':
                caseid_stx_app.caseid_account56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account57':
                caseid_stx_app.caseid_account57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account58':
                caseid_stx_app.caseid_account58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account59':
                caseid_stx_app.caseid_account59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account60':
                caseid_stx_app.caseid_account60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account61':
                caseid_stx_app.caseid_account61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account62':
                caseid_stx_app.caseid_account62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account63':
                caseid_stx_app.caseid_account63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account64':
                caseid_stx_app.caseid_account64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account65':
                caseid_stx_app.caseid_account65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account66':
                caseid_stx_app.caseid_account66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account67':
                caseid_stx_app.caseid_account67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account68':
                caseid_stx_app.caseid_account68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account69':
                caseid_stx_app.caseid_account69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account70':
                caseid_stx_app.caseid_account70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account71':
                caseid_stx_app.caseid_account71(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'App01':
                caseid_stx_app.caseid_app01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App02':
                caseid_stx_app.caseid_app02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App03':
                caseid_stx_app.caseid_app03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App04':
                caseid_stx_app.caseid_app04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App05':
                caseid_stx_app.caseid_app05(self)
        except:
            module_other_stx_app.close_app()



        try:
            if case == 'Home01':
                caseid_stx_app.caseid_home01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home02':
                caseid_stx_app.caseid_home02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home03':
                caseid_stx_app.caseid_home03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home04':
                caseid_stx_app.caseid_home04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home05':
                caseid_stx_app.caseid_home05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home06':
                caseid_stx_app.caseid_home06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home07':
                caseid_stx_app.caseid_home07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home08':
                caseid_stx_app.caseid_home08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09':
                caseid_stx_app.caseid_home09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home09_1':
                caseid_stx_app.caseid_home09_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09_2':
                caseid_stx_app.caseid_home09_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home10':
                caseid_stx_app.caseid_home10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home11':
                caseid_stx_app.caseid_home11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home12':
                caseid_stx_app.caseid_home12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home13':
                caseid_stx_app.caseid_home13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home14':
                caseid_stx_app.caseid_home14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home15':
                caseid_stx_app.caseid_home15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home16':
                caseid_stx_app.caseid_home16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home17':
                caseid_stx_app.caseid_home17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home18':
                caseid_stx_app.caseid_home18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home19':
                caseid_stx_app.caseid_home19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home20':
                caseid_stx_app.caseid_home20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home21':
                caseid_stx_app.caseid_home21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home22':
                caseid_stx_app.caseid_home22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home23':
                caseid_stx_app.caseid_home23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home24':
                caseid_stx_app.caseid_home24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home25':
                caseid_stx_app.caseid_home25(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home25_1':
                caseid_stx_app.caseid_home25_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home26':
                caseid_stx_app.caseid_home26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home27':
                caseid_stx_app.caseid_home27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home28':
                caseid_stx_app.caseid_home28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home29':
                caseid_stx_app.caseid_home29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30':
                caseid_stx_app.caseid_home30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30_1':
                caseid_stx_app.caseid_home30_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home31':
                caseid_stx_app.caseid_home31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home32':
                caseid_stx_app.caseid_home32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home33':
                caseid_stx_app.caseid_home33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home34':
                caseid_stx_app.caseid_home34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home35':
                caseid_stx_app.caseid_home35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home36':
                caseid_stx_app.caseid_home36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home37':
                caseid_stx_app.caseid_home37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home38':
                caseid_stx_app.caseid_home38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home39':
                caseid_stx_app.caseid_home39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home40':
                caseid_stx_app.caseid_home40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home41':
                caseid_stx_app.caseid_home41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42':
                caseid_stx_app.caseid_home42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_1':
                caseid_stx_app.caseid_home42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_2':
                caseid_stx_app.caseid_home42_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_3':
                caseid_stx_app.caseid_home42_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_4':
                caseid_stx_app.caseid_home42_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_5':
                caseid_stx_app.caseid_home42_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_6':
                caseid_stx_app.caseid_home42_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_7':
                caseid_stx_app.caseid_home42_7(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_8':
                caseid_stx_app.caseid_home42_8(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Home43':
                caseid_stx_app.caseid_home43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home43_1':
                caseid_stx_app.caseid_home43_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home44':
                caseid_stx_app.caseid_home44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home45':
                caseid_stx_app.caseid_home45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home46':
                caseid_stx_app.caseid_home46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47':
                caseid_stx_app.caseid_home47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_1':
                caseid_stx_app.caseid_home47_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_2':
                caseid_stx_app.caseid_home47_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_3':
                caseid_stx_app.caseid_home47_3(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Home48':
                caseid_stx_app.caseid_home48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home49':
                caseid_stx_app.caseid_home49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home50':
                caseid_stx_app.caseid_home50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home51':
                caseid_stx_app.caseid_home51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home52':
                caseid_stx_app.caseid_home52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home53':
                caseid_stx_app.caseid_home53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home54':
                caseid_stx_app.caseid_home54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home55':
                caseid_stx_app.caseid_home55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home56':
                caseid_stx_app.caseid_home56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home57':
                caseid_stx_app.caseid_home57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home58':
                caseid_stx_app.caseid_home58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home59':
                caseid_stx_app.caseid_home59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home60':
                caseid_stx_app.caseid_home60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home61':
                caseid_stx_app.caseid_home61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home62':
                caseid_stx_app.caseid_home62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home63':
                caseid_stx_app.caseid_home63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home64':
                caseid_stx_app.caseid_home64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home65':
                caseid_stx_app.caseid_home65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home66':
                caseid_stx_app.caseid_home66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home67':
                caseid_stx_app.caseid_home67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home68':
                caseid_stx_app.caseid_home68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home69':
                caseid_stx_app.caseid_home69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home70':
                caseid_stx_app.caseid_home70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home71':
                caseid_stx_app.caseid_home71(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home72':
                caseid_stx_app.caseid_home72(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home73':
                caseid_stx_app.caseid_home73(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home74':
                caseid_stx_app.caseid_home74(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home75':
                caseid_stx_app.caseid_home75(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home76':
                caseid_stx_app.caseid_home76(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home77':
                caseid_stx_app.caseid_home77(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home78':
                caseid_stx_app.caseid_home78(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home79':
                caseid_stx_app.caseid_home79(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home80':
                caseid_stx_app.caseid_home80(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home81':
                caseid_stx_app.caseid_home81(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home82':
                caseid_stx_app.caseid_home82(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Notification01':
                caseid_stx_app.caseid_notification01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification02':
                caseid_stx_app.caseid_notification02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification03':
                caseid_stx_app.caseid_notification03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification04':
                caseid_stx_app.caseid_notification04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification05':
                caseid_stx_app.caseid_notification05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification06':
                caseid_stx_app.caseid_notification06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification07':
                caseid_stx_app.caseid_notification07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification08':
                caseid_stx_app.caseid_notification08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification09':
                caseid_stx_app.caseid_notification09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification10':
                caseid_stx_app.caseid_notification10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'History01':
                caseid_stx_app.caseid_history01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02':
                caseid_stx_app.caseid_history02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02_1':
                caseid_stx_app.caseid_history02_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History03':
                caseid_stx_app.caseid_history03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History04':
                caseid_stx_app.caseid_history04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History05':
                caseid_stx_app.caseid_history05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History06':
                caseid_stx_app.caseid_history06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History07':
                caseid_stx_app.caseid_history07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History08':
                caseid_stx_app.caseid_history08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History09':
                caseid_stx_app.caseid_history09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History10':
                caseid_stx_app.caseid_history10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History11':
                caseid_stx_app.caseid_history11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History12':
                caseid_stx_app.caseid_history12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13':
                caseid_stx_app.caseid_history13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_1':
                caseid_stx_app.caseid_history13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_2':
                caseid_stx_app.caseid_history13_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_3':
                caseid_stx_app.caseid_history13_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_4':
                caseid_stx_app.caseid_history13_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_5':
                caseid_stx_app.caseid_history13_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History14':
                caseid_stx_app.caseid_history14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History15':
                caseid_stx_app.caseid_history15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History16':
                caseid_stx_app.caseid_history16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History17':
                caseid_stx_app.caseid_history17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History18':
                caseid_stx_app.caseid_history18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19':
                caseid_stx_app.caseid_history19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19_1':
                caseid_stx_app.caseid_history19_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History20':
                caseid_stx_app.caseid_history20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History21':
                caseid_stx_app.caseid_history21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History22':
                caseid_stx_app.caseid_history22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History23':
                caseid_stx_app.caseid_history23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History24':
                caseid_stx_app.caseid_history24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History25':
                caseid_stx_app.caseid_history25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History26':
                caseid_stx_app.caseid_history26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History27':
                caseid_stx_app.caseid_history27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History28':
                caseid_stx_app.caseid_history28(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'AccountWallet01':
                caseid_stx_app.caseid_accountwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet02':
                caseid_stx_app.caseid_accountwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet03':
                caseid_stx_app.caseid_accountwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet04':
                caseid_stx_app.caseid_accountwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet05':
                caseid_stx_app.caseid_accountwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet06':
                caseid_stx_app.caseid_accountwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet07':
                caseid_stx_app.caseid_accountwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet08':
                caseid_stx_app.caseid_accountwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet09':
                caseid_stx_app.caseid_accountwallet09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'CashWallet01':
                caseid_stx_app.caseid_cashwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet02':
                caseid_stx_app.caseid_cashwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet03':
                caseid_stx_app.caseid_cashwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet04':
                caseid_stx_app.caseid_cashwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet05':
                caseid_stx_app.caseid_cashwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet06':
                caseid_stx_app.caseid_cashwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet07':
                caseid_stx_app.caseid_cashwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet08':
                caseid_stx_app.caseid_cashwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet09':
                caseid_stx_app.caseid_cashwallet09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet10':
                caseid_stx_app.caseid_cashwallet10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'TripPayment01':
                caseid_stx_app.caseid_trippayment01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment02':
                caseid_stx_app.caseid_trippayment02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment03':
                caseid_stx_app.caseid_trippayment03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment04':
                caseid_stx_app.caseid_trippayment04(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'VnPayWallet01':
                caseid_stx_app.caseid_vnpaywallett01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'VnPayWallet02':
                caseid_stx_app.caseid_vnpaywallett02(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Setting01':
                caseid_stx_app.caseid_setting01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting02':
                caseid_stx_app.caseid_setting02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting03':
                caseid_stx_app.caseid_setting03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting04':
                caseid_stx_app.caseid_setting04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting05':
                caseid_stx_app.caseid_setting05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting06':
                caseid_stx_app.caseid_setting06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting07':
                caseid_stx_app.caseid_setting07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting08':
                caseid_stx_app.caseid_setting08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting09':
                caseid_stx_app.caseid_setting09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting10':
                caseid_stx_app.caseid_setting10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Report01':
                caseid_stx_app.caseid_report01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report02':
                caseid_stx_app.caseid_report02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report03':
                caseid_stx_app.caseid_report03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report04':
                caseid_stx_app.caseid_report04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report05':
                caseid_stx_app.caseid_report05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report06':
                caseid_stx_app.caseid_report06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report07':
                caseid_stx_app.caseid_report07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report08':
                caseid_stx_app.caseid_report08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report09':
                caseid_stx_app.caseid_report09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report10':
                caseid_stx_app.caseid_report10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report11':
                caseid_stx_app.caseid_report11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report12':
                caseid_stx_app.caseid_report12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report13':
                caseid_stx_app.caseid_report13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report14':
                caseid_stx_app.caseid_report14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report15':
                caseid_stx_app.caseid_report15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report16':
                caseid_stx_app.caseid_report16(self)
        except:
            module_other_stx_app.close_app()



        try:
            if case == 'Help01':
                caseid_stx_app.caseid_help01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help02':
                caseid_stx_app.caseid_help02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help03':
                caseid_stx_app.caseid_help03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help04':
                caseid_stx_app.caseid_help04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help05':
                caseid_stx_app.caseid_help05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Logout01':
                caseid_stx_app.caseid_logout01(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Web01':
                caseid_stx_app.caseid_web01(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web02':
                caseid_stx_app.caseid_web02(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web03':
                caseid_stx_app.caseid_web03(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web04':
                caseid_stx_app.caseid_web04(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web05':
                caseid_stx_app.caseid_web05(self)
        except:
            module_other_stx_app.close_chrome()





def retest_casefail(self):
    modetest = ''.join(re.findall(r'\d+', var_stx_app.modetest))
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        if modetest == "0":
            if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                print(sheet["A"+rownum].value)
                case_fail = sheet["A"+rownum].value
                list_casefail.append(case_fail)

        if modetest == "1":
            if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                print(sheet["A"+rownum].value)
                case_fail = sheet["A"+rownum].value
                list_casefail.append(case_fail)

        if modetest == "2":
            if sheet["J" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
                print(sheet["A" + rownum].value)
                case_fail = sheet["A" + rownum].value
                list_casefail.append(case_fail)

        if modetest == "3":
            if sheet["K" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
                print(sheet["A" + rownum].value)
                case_fail = sheet["A" + rownum].value
                list_casefail.append(case_fail)

        rownum = int(rownum)
    print(list_casefail)
    for case in list_casefail:
        try:
            if case == 'Login01':
                caseid_stx_app.caseid_login01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login02':
                caseid_stx_app.caseid_login02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login03':
                caseid_stx_app.caseid_login03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login04':
                caseid_stx_app.caseid_login04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login05':
                caseid_stx_app.caseid_login05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login06':
                caseid_stx_app.caseid_login06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login07':
                caseid_stx_app.caseid_login07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login08':
                caseid_stx_app.caseid_login08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login09':
                caseid_stx_app.caseid_login09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login10':
                caseid_stx_app.caseid_login10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'HomePage01':
                caseid_stx_app.caseid_homepage01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage02':
                caseid_stx_app.caseid_homepage02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage03':
                caseid_stx_app.caseid_homepage03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage04':
                caseid_stx_app.caseid_homepage04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage05':
                caseid_stx_app.caseid_homepage05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage06':
                caseid_stx_app.caseid_homepage06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage07':
                caseid_stx_app.caseid_homepage07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage08':
                caseid_stx_app.caseid_homepage08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage09':
                caseid_stx_app.caseid_homepage09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage10':
                caseid_stx_app.caseid_homepage10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage11':
                caseid_stx_app.caseid_homepage11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12':
                caseid_stx_app.caseid_homepage12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage12_1':
                caseid_stx_app.caseid_homepage12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12_2':
                caseid_stx_app.caseid_homepage12_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage13':
                caseid_stx_app.caseid_homepage13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage14':
                caseid_stx_app.caseid_homepage14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage15':
                caseid_stx_app.caseid_homepage15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage16':
                caseid_stx_app.caseid_homepage16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage17':
                caseid_stx_app.caseid_homepage17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage18':
                caseid_stx_app.caseid_homepage18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage19':
                caseid_stx_app.caseid_homepage19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage20':
                caseid_stx_app.caseid_homepage20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage21':
                caseid_stx_app.caseid_homepage21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage22':
                caseid_stx_app.caseid_homepage22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage23':
                caseid_stx_app.caseid_homepage23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage24':
                caseid_stx_app.caseid_homepage24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage25':
                caseid_stx_app.caseid_homepage25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage26':
                caseid_stx_app.caseid_homepage26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage27':
                caseid_stx_app.caseid_homepage27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28':
                caseid_stx_app.caseid_homepage28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28_1':
                caseid_stx_app.caseid_homepage28_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage29':
                caseid_stx_app.caseid_homepage29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage30':
                caseid_stx_app.caseid_homepage30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage31':
                caseid_stx_app.caseid_homepage31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage32':
                caseid_stx_app.caseid_homepage32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage33':
                caseid_stx_app.caseid_homepage33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage34':
                caseid_stx_app.caseid_homepage34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage35':
                caseid_stx_app.caseid_homepage35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage36':
                caseid_stx_app.caseid_homepage36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage37':
                caseid_stx_app.caseid_homepage37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage38':
                caseid_stx_app.caseid_homepage38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage39':
                caseid_stx_app.caseid_homepage39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage40':
                caseid_stx_app.caseid_homepage40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage41':
                caseid_stx_app.caseid_homepage41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42':
                caseid_stx_app.caseid_homepage42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42_1':
                caseid_stx_app.caseid_homepage42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43':
                caseid_stx_app.caseid_homepage43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_1':
                caseid_stx_app.caseid_homepage43_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_2':
                caseid_stx_app.caseid_homepage43_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage44':
                caseid_stx_app.caseid_homepage44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage45':
                caseid_stx_app.caseid_homepage45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage46':
                caseid_stx_app.caseid_homepage46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage47':
                caseid_stx_app.caseid_homepage47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage48':
                caseid_stx_app.caseid_homepage48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage49':
                caseid_stx_app.caseid_homepage49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage50':
                caseid_stx_app.caseid_homepage50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage51':
                caseid_stx_app.caseid_homepage51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage52':
                caseid_stx_app.caseid_homepage52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage53':
                caseid_stx_app.caseid_homepage53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage54':
                caseid_stx_app.caseid_homepage54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage55':
                caseid_stx_app.caseid_homepage55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage56':
                caseid_stx_app.caseid_homepage56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage57':
                caseid_stx_app.caseid_homepage57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage58':
                caseid_stx_app.caseid_homepage58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage59':
                caseid_stx_app.caseid_homepage59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage60':
                caseid_stx_app.caseid_homepage60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage61':
                caseid_stx_app.caseid_homepage61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage62':
                caseid_stx_app.caseid_homepage62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage63':
                caseid_stx_app.caseid_homepage63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage64':
                caseid_stx_app.caseid_homepage64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage65':
                caseid_stx_app.caseid_homepage65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage66':
                caseid_stx_app.caseid_homepage66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage67':
                caseid_stx_app.caseid_homepage67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage68':
                caseid_stx_app.caseid_homepage68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage69':
                caseid_stx_app.caseid_homepage69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage70':
                caseid_stx_app.caseid_homepage70(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'FavoriteSpot01':
                caseid_stx_app.caseid_favoritespot01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot02':
                caseid_stx_app.caseid_favoritespot02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot03':
                caseid_stx_app.caseid_favoritespot03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot04':
                caseid_stx_app.caseid_favoritespot04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot05':
                caseid_stx_app.caseid_favoritespot05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot06':
                caseid_stx_app.caseid_favoritespot06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot07':
                caseid_stx_app.caseid_favoritespot07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot08':
                caseid_stx_app.caseid_favoritespot08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot09':
                caseid_stx_app.caseid_favoritespot09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot10':
                caseid_stx_app.caseid_favoritespot10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot11':
                caseid_stx_app.caseid_favoritespot11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot12':
                caseid_stx_app.caseid_favoritespot12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Promotion01':
                caseid_stx_app.caseid_promotion01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion02':
                caseid_stx_app.caseid_promotion02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion03':
                caseid_stx_app.caseid_promotion03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion04':
                caseid_stx_app.caseid_promotion04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion05':
                caseid_stx_app.caseid_promotion05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion06':
                caseid_stx_app.caseid_promotion06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion07':
                caseid_stx_app.caseid_promotion07(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account01':
                caseid_stx_app.caseid_account01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account02':
                caseid_stx_app.caseid_account02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account03':
                caseid_stx_app.caseid_account03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account04':
                caseid_stx_app.caseid_account04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account05':
                caseid_stx_app.caseid_account05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account06':
                caseid_stx_app.caseid_account06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account07':
                caseid_stx_app.caseid_account07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account08':
                caseid_stx_app.caseid_account08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account09':
                caseid_stx_app.caseid_account09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account10':
                caseid_stx_app.caseid_account10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account11':
                caseid_stx_app.caseid_account11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12':
                caseid_stx_app.caseid_account12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_1':
                caseid_stx_app.caseid_account12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_2':
                caseid_stx_app.caseid_account12_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_3':
                caseid_stx_app.caseid_account12_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_4':
                caseid_stx_app.caseid_account12_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_5':
                caseid_stx_app.caseid_account12_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_6':
                caseid_stx_app.caseid_account12_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13':
                caseid_stx_app.caseid_account13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13_1':
                caseid_stx_app.caseid_account13_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account14':
                caseid_stx_app.caseid_account14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account15':
                caseid_stx_app.caseid_account15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account16':
                caseid_stx_app.caseid_account16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account17':
                caseid_stx_app.caseid_account17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account18':
                caseid_stx_app.caseid_account18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account19':
                caseid_stx_app.caseid_account19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account20':
                caseid_stx_app.caseid_account20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account21':
                caseid_stx_app.caseid_account21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account22':
                caseid_stx_app.caseid_account22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account23':
                caseid_stx_app.caseid_account23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account24':
                caseid_stx_app.caseid_account24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account25':
                caseid_stx_app.caseid_account25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account26':
                caseid_stx_app.caseid_account26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account27':
                caseid_stx_app.caseid_account27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account28':
                caseid_stx_app.caseid_account28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account29':
                caseid_stx_app.caseid_account29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account30':
                caseid_stx_app.caseid_account30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account31':
                caseid_stx_app.caseid_account31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account32':
                caseid_stx_app.caseid_account32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account33':
                caseid_stx_app.caseid_account33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account34':
                caseid_stx_app.caseid_account34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account35':
                caseid_stx_app.caseid_account35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account36':
                caseid_stx_app.caseid_account36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account37':
                caseid_stx_app.caseid_account37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account38':
                caseid_stx_app.caseid_account38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account39':
                caseid_stx_app.caseid_account39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40':
                caseid_stx_app.caseid_account40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40_1':
                caseid_stx_app.caseid_account40_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account41':
                caseid_stx_app.caseid_account41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account42':
                caseid_stx_app.caseid_account42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account43':
                caseid_stx_app.caseid_account43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account44':
                caseid_stx_app.caseid_account44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account45':
                caseid_stx_app.caseid_account45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account46':
                caseid_stx_app.caseid_account46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account47':
                caseid_stx_app.caseid_account47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account48':
                caseid_stx_app.caseid_account48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account49':
                caseid_stx_app.caseid_account49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50':
                caseid_stx_app.caseid_account50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50_1':
                caseid_stx_app.caseid_account50_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51':
                caseid_stx_app.caseid_account51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_1':
                caseid_stx_app.caseid_account51_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_2':
                caseid_stx_app.caseid_account51_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account52':
                caseid_stx_app.caseid_account52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account53':
                caseid_stx_app.caseid_account53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account54':
                caseid_stx_app.caseid_account54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account55':
                caseid_stx_app.caseid_account55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account56':
                caseid_stx_app.caseid_account56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account57':
                caseid_stx_app.caseid_account57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account58':
                caseid_stx_app.caseid_account58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account59':
                caseid_stx_app.caseid_account59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account60':
                caseid_stx_app.caseid_account60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account61':
                caseid_stx_app.caseid_account61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account62':
                caseid_stx_app.caseid_account62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account63':
                caseid_stx_app.caseid_account63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account64':
                caseid_stx_app.caseid_account64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account65':
                caseid_stx_app.caseid_account65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account66':
                caseid_stx_app.caseid_account66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account67':
                caseid_stx_app.caseid_account67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account68':
                caseid_stx_app.caseid_account68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account69':
                caseid_stx_app.caseid_account69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account70':
                caseid_stx_app.caseid_account70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account71':
                caseid_stx_app.caseid_account71(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'App01':
                caseid_stx_app.caseid_app01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App02':
                caseid_stx_app.caseid_app02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App03':
                caseid_stx_app.caseid_app03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App04':
                caseid_stx_app.caseid_app04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App05':
                caseid_stx_app.caseid_app05(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Home01':
                caseid_stx_app.caseid_home01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home02':
                caseid_stx_app.caseid_home02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home03':
                caseid_stx_app.caseid_home03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home04':
                caseid_stx_app.caseid_home04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home05':
                caseid_stx_app.caseid_home05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home06':
                caseid_stx_app.caseid_home06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home07':
                caseid_stx_app.caseid_home07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home08':
                caseid_stx_app.caseid_home08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09':
                caseid_stx_app.caseid_home09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home09_1':
                caseid_stx_app.caseid_home09_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09_2':
                caseid_stx_app.caseid_home09_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home10':
                caseid_stx_app.caseid_home10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home11':
                caseid_stx_app.caseid_home11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home12':
                caseid_stx_app.caseid_home12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home13':
                caseid_stx_app.caseid_home13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home14':
                caseid_stx_app.caseid_home14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home15':
                caseid_stx_app.caseid_home15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home16':
                caseid_stx_app.caseid_home16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home17':
                caseid_stx_app.caseid_home17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home18':
                caseid_stx_app.caseid_home18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home19':
                caseid_stx_app.caseid_home19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home20':
                caseid_stx_app.caseid_home20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home21':
                caseid_stx_app.caseid_home21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home22':
                caseid_stx_app.caseid_home22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home23':
                caseid_stx_app.caseid_home23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home24':
                caseid_stx_app.caseid_home24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home25':
                caseid_stx_app.caseid_home25(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home25_1':
                caseid_stx_app.caseid_home25_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home26':
                caseid_stx_app.caseid_home26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home27':
                caseid_stx_app.caseid_home27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home28':
                caseid_stx_app.caseid_home28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home29':
                caseid_stx_app.caseid_home29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30':
                caseid_stx_app.caseid_home30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30_1':
                caseid_stx_app.caseid_home30_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home31':
                caseid_stx_app.caseid_home31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home32':
                caseid_stx_app.caseid_home32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home33':
                caseid_stx_app.caseid_home33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home34':
                caseid_stx_app.caseid_home34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home35':
                caseid_stx_app.caseid_home35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home36':
                caseid_stx_app.caseid_home36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home37':
                caseid_stx_app.caseid_home37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home38':
                caseid_stx_app.caseid_home38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home39':
                caseid_stx_app.caseid_home39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home40':
                caseid_stx_app.caseid_home40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home41':
                caseid_stx_app.caseid_home41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42':
                caseid_stx_app.caseid_home42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_1':
                caseid_stx_app.caseid_home42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_2':
                caseid_stx_app.caseid_home42_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_3':
                caseid_stx_app.caseid_home42_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_4':
                caseid_stx_app.caseid_home42_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_5':
                caseid_stx_app.caseid_home42_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_6':
                caseid_stx_app.caseid_home42_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_7':
                caseid_stx_app.caseid_home42_7(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_8':
                caseid_stx_app.caseid_home42_8(self)
        except:
            module_other_stx_app.close_app()



        try:
            if case == 'Home43':
                caseid_stx_app.caseid_home43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home43_1':
                caseid_stx_app.caseid_home43_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home44':
                caseid_stx_app.caseid_home44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home45':
                caseid_stx_app.caseid_home45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home46':
                caseid_stx_app.caseid_home46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47':
                caseid_stx_app.caseid_home47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_1':
                caseid_stx_app.caseid_home47_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_2':
                caseid_stx_app.caseid_home47_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_3':
                caseid_stx_app.caseid_home47_3(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home48':
                caseid_stx_app.caseid_home48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home49':
                caseid_stx_app.caseid_home49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home50':
                caseid_stx_app.caseid_home50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home51':
                caseid_stx_app.caseid_home51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home52':
                caseid_stx_app.caseid_home52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home53':
                caseid_stx_app.caseid_home53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home54':
                caseid_stx_app.caseid_home54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home55':
                caseid_stx_app.caseid_home55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home56':
                caseid_stx_app.caseid_home56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home57':
                caseid_stx_app.caseid_home57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home58':
                caseid_stx_app.caseid_home58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home59':
                caseid_stx_app.caseid_home59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home60':
                caseid_stx_app.caseid_home60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home61':
                caseid_stx_app.caseid_home61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home62':
                caseid_stx_app.caseid_home62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home63':
                caseid_stx_app.caseid_home63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home64':
                caseid_stx_app.caseid_home64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home65':
                caseid_stx_app.caseid_home65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home66':
                caseid_stx_app.caseid_home66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home67':
                caseid_stx_app.caseid_home67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home68':
                caseid_stx_app.caseid_home68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home69':
                caseid_stx_app.caseid_home69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home70':
                caseid_stx_app.caseid_home70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home71':
                caseid_stx_app.caseid_home71(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home72':
                caseid_stx_app.caseid_home72(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home73':
                caseid_stx_app.caseid_home73(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home74':
                caseid_stx_app.caseid_home74(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home75':
                caseid_stx_app.caseid_home75(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home76':
                caseid_stx_app.caseid_home76(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home77':
                caseid_stx_app.caseid_home77(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home78':
                caseid_stx_app.caseid_home78(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home79':
                caseid_stx_app.caseid_home79(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home80':
                caseid_stx_app.caseid_home80(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home81':
                caseid_stx_app.caseid_home81(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home82':
                caseid_stx_app.caseid_home82(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Notification01':
                caseid_stx_app.caseid_notification01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification02':
                caseid_stx_app.caseid_notification02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification03':
                caseid_stx_app.caseid_notification03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification04':
                caseid_stx_app.caseid_notification04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification05':
                caseid_stx_app.caseid_notification05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification06':
                caseid_stx_app.caseid_notification06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification07':
                caseid_stx_app.caseid_notification07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification08':
                caseid_stx_app.caseid_notification08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification09':
                caseid_stx_app.caseid_notification09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification10':
                caseid_stx_app.caseid_notification10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'History01':
                caseid_stx_app.caseid_history01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02':
                caseid_stx_app.caseid_history02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02_1':
                caseid_stx_app.caseid_history02_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History03':
                caseid_stx_app.caseid_history03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History04':
                caseid_stx_app.caseid_history04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History05':
                caseid_stx_app.caseid_history05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History06':
                caseid_stx_app.caseid_history06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History07':
                caseid_stx_app.caseid_history07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History08':
                caseid_stx_app.caseid_history08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History09':
                caseid_stx_app.caseid_history09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History10':
                caseid_stx_app.caseid_history10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History11':
                caseid_stx_app.caseid_history11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History12':
                caseid_stx_app.caseid_history12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13':
                caseid_stx_app.caseid_history13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_1':
                caseid_stx_app.caseid_history13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_2':
                caseid_stx_app.caseid_history13_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_3':
                caseid_stx_app.caseid_history13_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_4':
                caseid_stx_app.caseid_history13_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_5':
                caseid_stx_app.caseid_history13_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History14':
                caseid_stx_app.caseid_history14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History15':
                caseid_stx_app.caseid_history15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History16':
                caseid_stx_app.caseid_history16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History17':
                caseid_stx_app.caseid_history17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History18':
                caseid_stx_app.caseid_history18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19':
                caseid_stx_app.caseid_history19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19_1':
                caseid_stx_app.caseid_history19_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History20':
                caseid_stx_app.caseid_history20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History21':
                caseid_stx_app.caseid_history21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History22':
                caseid_stx_app.caseid_history22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History23':
                caseid_stx_app.caseid_history23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History24':
                caseid_stx_app.caseid_history24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History25':
                caseid_stx_app.caseid_history25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History26':
                caseid_stx_app.caseid_history26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History27':
                caseid_stx_app.caseid_history27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History28':
                caseid_stx_app.caseid_history28(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'AccountWallet01':
                caseid_stx_app.caseid_accountwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet02':
                caseid_stx_app.caseid_accountwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet03':
                caseid_stx_app.caseid_accountwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet04':
                caseid_stx_app.caseid_accountwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet05':
                caseid_stx_app.caseid_accountwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet06':
                caseid_stx_app.caseid_accountwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet07':
                caseid_stx_app.caseid_accountwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet08':
                caseid_stx_app.caseid_accountwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet09':
                caseid_stx_app.caseid_accountwallet09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'CashWallet01':
                caseid_stx_app.caseid_cashwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet02':
                caseid_stx_app.caseid_cashwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet03':
                caseid_stx_app.caseid_cashwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet04':
                caseid_stx_app.caseid_cashwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet05':
                caseid_stx_app.caseid_cashwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet06':
                caseid_stx_app.caseid_cashwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet07':
                caseid_stx_app.caseid_cashwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet08':
                caseid_stx_app.caseid_cashwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet09':
                caseid_stx_app.caseid_cashwallet09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet10':
                caseid_stx_app.caseid_cashwallet10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'TripPayment01':
                caseid_stx_app.caseid_trippayment01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment02':
                caseid_stx_app.caseid_trippayment02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment03':
                caseid_stx_app.caseid_trippayment03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment04':
                caseid_stx_app.caseid_trippayment04(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'VnPayWallet01':
                caseid_stx_app.caseid_vnpaywallett01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'VnPayWallet02':
                caseid_stx_app.caseid_vnpaywallett02(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Setting01':
                caseid_stx_app.caseid_setting01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting02':
                caseid_stx_app.caseid_setting02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting03':
                caseid_stx_app.caseid_setting03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting04':
                caseid_stx_app.caseid_setting04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting05':
                caseid_stx_app.caseid_setting05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting06':
                caseid_stx_app.caseid_setting06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting07':
                caseid_stx_app.caseid_setting07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting08':
                caseid_stx_app.caseid_setting08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting09':
                caseid_stx_app.caseid_setting09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting10':
                caseid_stx_app.caseid_setting10(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Report01':
                caseid_stx_app.caseid_report01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report02':
                caseid_stx_app.caseid_report02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report03':
                caseid_stx_app.caseid_report03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report04':
                caseid_stx_app.caseid_report04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report05':
                caseid_stx_app.caseid_report05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report06':
                caseid_stx_app.caseid_report06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report07':
                caseid_stx_app.caseid_report07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report08':
                caseid_stx_app.caseid_report08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report09':
                caseid_stx_app.caseid_report09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report10':
                caseid_stx_app.caseid_report10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report11':
                caseid_stx_app.caseid_report11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report12':
                caseid_stx_app.caseid_report12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report13':
                caseid_stx_app.caseid_report13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report14':
                caseid_stx_app.caseid_report14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report15':
                caseid_stx_app.caseid_report15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report16':
                caseid_stx_app.caseid_report16(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Help01':
                caseid_stx_app.caseid_help01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help02':
                caseid_stx_app.caseid_help02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help03':
                caseid_stx_app.caseid_help03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help04':
                caseid_stx_app.caseid_help04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help05':
                caseid_stx_app.caseid_help05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Logout01':
                caseid_stx_app.caseid_logout01(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Web01':
                caseid_stx_app.caseid_web01(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web02':
                caseid_stx_app.caseid_web02(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web03':
                caseid_stx_app.caseid_web03(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web04':
                caseid_stx_app.caseid_web04(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web05':
                caseid_stx_app.caseid_web05(self)
        except:
            module_other_stx_app.close_chrome()





def luong_all(self):
    list_0 = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        print(sheet["G" + rownum].value)
        print(sheet["H" + rownum].value)
        if (sheet["H" + rownum].value == "x") and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case = sheet["A" + rownum].value
            list_0.append(case)
        rownum = int(rownum)
    print(list_0)
    print("Tổng số case luồng full: ", len(list_0))
    for case in list_0:
        try:
            if case == 'Login01':
                caseid_stx_app.caseid_login01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login02':
                caseid_stx_app.caseid_login02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login03':
                caseid_stx_app.caseid_login03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login04':
                caseid_stx_app.caseid_login04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login05':
                caseid_stx_app.caseid_login05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login06':
                caseid_stx_app.caseid_login06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login07':
                caseid_stx_app.caseid_login07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login08':
                caseid_stx_app.caseid_login08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login09':
                caseid_stx_app.caseid_login09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login10':
                caseid_stx_app.caseid_login10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage01':
                caseid_stx_app.caseid_homepage01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage02':
                caseid_stx_app.caseid_homepage02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage03':
                caseid_stx_app.caseid_homepage03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage04':
                caseid_stx_app.caseid_homepage04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage05':
                caseid_stx_app.caseid_homepage05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage06':
                caseid_stx_app.caseid_homepage06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage07':
                caseid_stx_app.caseid_homepage07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage08':
                caseid_stx_app.caseid_homepage08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage09':
                caseid_stx_app.caseid_homepage09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage10':
                caseid_stx_app.caseid_homepage10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage11':
                caseid_stx_app.caseid_homepage11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12':
                caseid_stx_app.caseid_homepage12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage12_1':
                caseid_stx_app.caseid_homepage12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12_2':
                caseid_stx_app.caseid_homepage12_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage13':
                caseid_stx_app.caseid_homepage13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage14':
                caseid_stx_app.caseid_homepage14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage15':
                caseid_stx_app.caseid_homepage15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage16':
                caseid_stx_app.caseid_homepage16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage17':
                caseid_stx_app.caseid_homepage17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage18':
                caseid_stx_app.caseid_homepage18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage19':
                caseid_stx_app.caseid_homepage19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage20':
                caseid_stx_app.caseid_homepage20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage21':
                caseid_stx_app.caseid_homepage21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage22':
                caseid_stx_app.caseid_homepage22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage23':
                caseid_stx_app.caseid_homepage23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage24':
                caseid_stx_app.caseid_homepage24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage25':
                caseid_stx_app.caseid_homepage25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage26':
                caseid_stx_app.caseid_homepage26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage27':
                caseid_stx_app.caseid_homepage27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28':
                caseid_stx_app.caseid_homepage28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28_1':
                caseid_stx_app.caseid_homepage28_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage29':
                caseid_stx_app.caseid_homepage29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage30':
                caseid_stx_app.caseid_homepage30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage31':
                caseid_stx_app.caseid_homepage31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage32':
                caseid_stx_app.caseid_homepage32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage33':
                caseid_stx_app.caseid_homepage33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage34':
                caseid_stx_app.caseid_homepage34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage35':
                caseid_stx_app.caseid_homepage35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage36':
                caseid_stx_app.caseid_homepage36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage37':
                caseid_stx_app.caseid_homepage37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage38':
                caseid_stx_app.caseid_homepage38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage39':
                caseid_stx_app.caseid_homepage39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage40':
                caseid_stx_app.caseid_homepage40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage41':
                caseid_stx_app.caseid_homepage41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42':
                caseid_stx_app.caseid_homepage42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42_1':
                caseid_stx_app.caseid_homepage42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43':
                caseid_stx_app.caseid_homepage43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_1':
                caseid_stx_app.caseid_homepage43_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_2':
                caseid_stx_app.caseid_homepage43_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage44':
                caseid_stx_app.caseid_homepage44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage45':
                caseid_stx_app.caseid_homepage45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage46':
                caseid_stx_app.caseid_homepage46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage47':
                caseid_stx_app.caseid_homepage47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage48':
                caseid_stx_app.caseid_homepage48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage49':
                caseid_stx_app.caseid_homepage49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage50':
                caseid_stx_app.caseid_homepage50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage51':
                caseid_stx_app.caseid_homepage51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage52':
                caseid_stx_app.caseid_homepage52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage53':
                caseid_stx_app.caseid_homepage53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage54':
                caseid_stx_app.caseid_homepage54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage55':
                caseid_stx_app.caseid_homepage55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage56':
                caseid_stx_app.caseid_homepage56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage57':
                caseid_stx_app.caseid_homepage57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage58':
                caseid_stx_app.caseid_homepage58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage59':
                caseid_stx_app.caseid_homepage59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage60':
                caseid_stx_app.caseid_homepage60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage61':
                caseid_stx_app.caseid_homepage61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage62':
                caseid_stx_app.caseid_homepage62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage63':
                caseid_stx_app.caseid_homepage63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage64':
                caseid_stx_app.caseid_homepage64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage65':
                caseid_stx_app.caseid_homepage65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage66':
                caseid_stx_app.caseid_homepage66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage67':
                caseid_stx_app.caseid_homepage67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage68':
                caseid_stx_app.caseid_homepage68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage69':
                caseid_stx_app.caseid_homepage69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage70':
                caseid_stx_app.caseid_homepage70(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'FavoriteSpot01':
                caseid_stx_app.caseid_favoritespot01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot02':
                caseid_stx_app.caseid_favoritespot02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot03':
                caseid_stx_app.caseid_favoritespot03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot04':
                caseid_stx_app.caseid_favoritespot04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot05':
                caseid_stx_app.caseid_favoritespot05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot06':
                caseid_stx_app.caseid_favoritespot06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot07':
                caseid_stx_app.caseid_favoritespot07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot08':
                caseid_stx_app.caseid_favoritespot08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot09':
                caseid_stx_app.caseid_favoritespot09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot10':
                caseid_stx_app.caseid_favoritespot10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot11':
                caseid_stx_app.caseid_favoritespot11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot12':
                caseid_stx_app.caseid_favoritespot12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Promotion01':
                caseid_stx_app.caseid_promotion01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion02':
                caseid_stx_app.caseid_promotion02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion03':
                caseid_stx_app.caseid_promotion03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion04':
                caseid_stx_app.caseid_promotion04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion05':
                caseid_stx_app.caseid_promotion05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion06':
                caseid_stx_app.caseid_promotion06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion07':
                caseid_stx_app.caseid_promotion07(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account01':
                caseid_stx_app.caseid_account01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account02':
                caseid_stx_app.caseid_account02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account03':
                caseid_stx_app.caseid_account03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account04':
                caseid_stx_app.caseid_account04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account05':
                caseid_stx_app.caseid_account05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account06':
                caseid_stx_app.caseid_account06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account07':
                caseid_stx_app.caseid_account07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account08':
                caseid_stx_app.caseid_account08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account09':
                caseid_stx_app.caseid_account09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account10':
                caseid_stx_app.caseid_account10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account11':
                caseid_stx_app.caseid_account11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12':
                caseid_stx_app.caseid_account12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_1':
                caseid_stx_app.caseid_account12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_2':
                caseid_stx_app.caseid_account12_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_3':
                caseid_stx_app.caseid_account12_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_4':
                caseid_stx_app.caseid_account12_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_5':
                caseid_stx_app.caseid_account12_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_6':
                caseid_stx_app.caseid_account12_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13':
                caseid_stx_app.caseid_account13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13_1':
                caseid_stx_app.caseid_account13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account14':
                caseid_stx_app.caseid_account14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account15':
                caseid_stx_app.caseid_account15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account16':
                caseid_stx_app.caseid_account16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account17':
                caseid_stx_app.caseid_account17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account18':
                caseid_stx_app.caseid_account18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account19':
                caseid_stx_app.caseid_account19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account20':
                caseid_stx_app.caseid_account20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account21':
                caseid_stx_app.caseid_account21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account22':
                caseid_stx_app.caseid_account22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account23':
                caseid_stx_app.caseid_account23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account24':
                caseid_stx_app.caseid_account24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account25':
                caseid_stx_app.caseid_account25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account26':
                caseid_stx_app.caseid_account26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account27':
                caseid_stx_app.caseid_account27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account28':
                caseid_stx_app.caseid_account28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account29':
                caseid_stx_app.caseid_account29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account30':
                caseid_stx_app.caseid_account30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account31':
                caseid_stx_app.caseid_account31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account32':
                caseid_stx_app.caseid_account32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account33':
                caseid_stx_app.caseid_account33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account34':
                caseid_stx_app.caseid_account34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account35':
                caseid_stx_app.caseid_account35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account36':
                caseid_stx_app.caseid_account36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account37':
                caseid_stx_app.caseid_account37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account38':
                caseid_stx_app.caseid_account38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account39':
                caseid_stx_app.caseid_account39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40':
                caseid_stx_app.caseid_account40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40_1':
                caseid_stx_app.caseid_account40_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account41':
                caseid_stx_app.caseid_account41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account42':
                caseid_stx_app.caseid_account42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account43':
                caseid_stx_app.caseid_account43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account44':
                caseid_stx_app.caseid_account44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account45':
                caseid_stx_app.caseid_account45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account46':
                caseid_stx_app.caseid_account46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account47':
                caseid_stx_app.caseid_account47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account48':
                caseid_stx_app.caseid_account48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account49':
                caseid_stx_app.caseid_account49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50':
                caseid_stx_app.caseid_account50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50_1':
                caseid_stx_app.caseid_account50_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51':
                caseid_stx_app.caseid_account51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_1':
                caseid_stx_app.caseid_account51_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_2':
                caseid_stx_app.caseid_account51_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account52':
                caseid_stx_app.caseid_account52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account53':
                caseid_stx_app.caseid_account53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account54':
                caseid_stx_app.caseid_account54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account55':
                caseid_stx_app.caseid_account55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account56':
                caseid_stx_app.caseid_account56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account57':
                caseid_stx_app.caseid_account57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account58':
                caseid_stx_app.caseid_account58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account59':
                caseid_stx_app.caseid_account59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account60':
                caseid_stx_app.caseid_account60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account61':
                caseid_stx_app.caseid_account61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account62':
                caseid_stx_app.caseid_account62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account63':
                caseid_stx_app.caseid_account63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account64':
                caseid_stx_app.caseid_account64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account65':
                caseid_stx_app.caseid_account65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account66':
                caseid_stx_app.caseid_account66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account67':
                caseid_stx_app.caseid_account67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account68':
                caseid_stx_app.caseid_account68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account69':
                caseid_stx_app.caseid_account69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account70':
                caseid_stx_app.caseid_account70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account71':
                caseid_stx_app.caseid_account71(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'App01':
                caseid_stx_app.caseid_app01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App02':
                caseid_stx_app.caseid_app02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App03':
                caseid_stx_app.caseid_app03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App04':
                caseid_stx_app.caseid_app04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App05':
                caseid_stx_app.caseid_app05(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home01':
                caseid_stx_app.caseid_home01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home02':
                caseid_stx_app.caseid_home02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home03':
                caseid_stx_app.caseid_home03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home04':
                caseid_stx_app.caseid_home04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home05':
                caseid_stx_app.caseid_home05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home06':
                caseid_stx_app.caseid_home06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home07':
                caseid_stx_app.caseid_home07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home08':
                caseid_stx_app.caseid_home08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09':
                caseid_stx_app.caseid_home09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home09_1':
                caseid_stx_app.caseid_home09_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09_2':
                caseid_stx_app.caseid_home09_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home10':
                caseid_stx_app.caseid_home10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home11':
                caseid_stx_app.caseid_home11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home12':
                caseid_stx_app.caseid_home12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home13':
                caseid_stx_app.caseid_home13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home14':
                caseid_stx_app.caseid_home14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home15':
                caseid_stx_app.caseid_home15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home16':
                caseid_stx_app.caseid_home16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home17':
                caseid_stx_app.caseid_home17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home18':
                caseid_stx_app.caseid_home18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home19':
                caseid_stx_app.caseid_home19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home20':
                caseid_stx_app.caseid_home20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home21':
                caseid_stx_app.caseid_home21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home22':
                caseid_stx_app.caseid_home22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home23':
                caseid_stx_app.caseid_home23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home24':
                caseid_stx_app.caseid_home24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home25':
                caseid_stx_app.caseid_home25(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home25_1':
                caseid_stx_app.caseid_home25_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home26':
                caseid_stx_app.caseid_home26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home27':
                caseid_stx_app.caseid_home27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home28':
                caseid_stx_app.caseid_home28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home29':
                caseid_stx_app.caseid_home29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30':
                caseid_stx_app.caseid_home30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30_1':
                caseid_stx_app.caseid_home30_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home31':
                caseid_stx_app.caseid_home31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home32':
                caseid_stx_app.caseid_home32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home33':
                caseid_stx_app.caseid_home33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home34':
                caseid_stx_app.caseid_home34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home35':
                caseid_stx_app.caseid_home35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home36':
                caseid_stx_app.caseid_home36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home37':
                caseid_stx_app.caseid_home37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home38':
                caseid_stx_app.caseid_home38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home39':
                caseid_stx_app.caseid_home39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home40':
                caseid_stx_app.caseid_home40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home41':
                caseid_stx_app.caseid_home41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42':
                caseid_stx_app.caseid_home42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_1':
                caseid_stx_app.caseid_home42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_2':
                caseid_stx_app.caseid_home42_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_3':
                caseid_stx_app.caseid_home42_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_4':
                caseid_stx_app.caseid_home42_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_5':
                caseid_stx_app.caseid_home42_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_6':
                caseid_stx_app.caseid_home42_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_7':
                caseid_stx_app.caseid_home42_7(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_8':
                caseid_stx_app.caseid_home42_8(self)
        except:
            module_other_stx_app.close_app()



        try:
            if case == 'Home43':
                caseid_stx_app.caseid_home43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home43_1':
                caseid_stx_app.caseid_home43_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home44':
                caseid_stx_app.caseid_home44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home45':
                caseid_stx_app.caseid_home45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home46':
                caseid_stx_app.caseid_home46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47':
                caseid_stx_app.caseid_home47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_1':
                caseid_stx_app.caseid_home47_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_2':
                caseid_stx_app.caseid_home47_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_3':
                caseid_stx_app.caseid_home47_3(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home48':
                caseid_stx_app.caseid_home48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home49':
                caseid_stx_app.caseid_home49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home50':
                caseid_stx_app.caseid_home50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home51':
                caseid_stx_app.caseid_home51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home52':
                caseid_stx_app.caseid_home52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home53':
                caseid_stx_app.caseid_home53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home54':
                caseid_stx_app.caseid_home54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home55':
                caseid_stx_app.caseid_home55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home56':
                caseid_stx_app.caseid_home56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home57':
                caseid_stx_app.caseid_home57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home58':
                caseid_stx_app.caseid_home58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home59':
                caseid_stx_app.caseid_home59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home60':
                caseid_stx_app.caseid_home60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home61':
                caseid_stx_app.caseid_home61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home62':
                caseid_stx_app.caseid_home62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home63':
                caseid_stx_app.caseid_home63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home64':
                caseid_stx_app.caseid_home64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home65':
                caseid_stx_app.caseid_home65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home66':
                caseid_stx_app.caseid_home66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home67':
                caseid_stx_app.caseid_home67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home68':
                caseid_stx_app.caseid_home68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home69':
                caseid_stx_app.caseid_home69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home70':
                caseid_stx_app.caseid_home70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home71':
                caseid_stx_app.caseid_home71(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home72':
                caseid_stx_app.caseid_home72(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home73':
                caseid_stx_app.caseid_home73(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home74':
                caseid_stx_app.caseid_home74(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home75':
                caseid_stx_app.caseid_home75(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home76':
                caseid_stx_app.caseid_home76(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home77':
                caseid_stx_app.caseid_home77(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home78':
                caseid_stx_app.caseid_home78(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home79':
                caseid_stx_app.caseid_home79(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home80':
                caseid_stx_app.caseid_home80(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home81':
                caseid_stx_app.caseid_home81(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home82':
                caseid_stx_app.caseid_home82(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Notification01':
                caseid_stx_app.caseid_notification01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification02':
                caseid_stx_app.caseid_notification02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification03':
                caseid_stx_app.caseid_notification03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification04':
                caseid_stx_app.caseid_notification04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification05':
                caseid_stx_app.caseid_notification05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification06':
                caseid_stx_app.caseid_notification06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification07':
                caseid_stx_app.caseid_notification07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification08':
                caseid_stx_app.caseid_notification08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification09':
                caseid_stx_app.caseid_notification09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification10':
                caseid_stx_app.caseid_notification10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History01':
                caseid_stx_app.caseid_history01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02':
                caseid_stx_app.caseid_history02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02_1':
                caseid_stx_app.caseid_history02_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History03':
                caseid_stx_app.caseid_history03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History04':
                caseid_stx_app.caseid_history04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History05':
                caseid_stx_app.caseid_history05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History06':
                caseid_stx_app.caseid_history06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History07':
                caseid_stx_app.caseid_history07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History08':
                caseid_stx_app.caseid_history08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History09':
                caseid_stx_app.caseid_history09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History10':
                caseid_stx_app.caseid_history10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History11':
                caseid_stx_app.caseid_history11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History12':
                caseid_stx_app.caseid_history12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13':
                caseid_stx_app.caseid_history13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_1':
                caseid_stx_app.caseid_history13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_2':
                caseid_stx_app.caseid_history13_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_3':
                caseid_stx_app.caseid_history13_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_4':
                caseid_stx_app.caseid_history13_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_5':
                caseid_stx_app.caseid_history13_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History14':
                caseid_stx_app.caseid_history14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History15':
                caseid_stx_app.caseid_history15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History16':
                caseid_stx_app.caseid_history16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History17':
                caseid_stx_app.caseid_history17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History18':
                caseid_stx_app.caseid_history18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19':
                caseid_stx_app.caseid_history19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19_1':
                caseid_stx_app.caseid_history19_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History20':
                caseid_stx_app.caseid_history20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History21':
                caseid_stx_app.caseid_history21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History22':
                caseid_stx_app.caseid_history22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History23':
                caseid_stx_app.caseid_history23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History24':
                caseid_stx_app.caseid_history24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History25':
                caseid_stx_app.caseid_history25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History26':
                caseid_stx_app.caseid_history26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History27':
                caseid_stx_app.caseid_history27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History28':
                caseid_stx_app.caseid_history28(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'AccountWallet01':
                caseid_stx_app.caseid_accountwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet02':
                caseid_stx_app.caseid_accountwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet03':
                caseid_stx_app.caseid_accountwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet04':
                caseid_stx_app.caseid_accountwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet05':
                caseid_stx_app.caseid_accountwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet06':
                caseid_stx_app.caseid_accountwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet07':
                caseid_stx_app.caseid_accountwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet08':
                caseid_stx_app.caseid_accountwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet09':
                caseid_stx_app.caseid_accountwallet09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'CashWallet01':
                caseid_stx_app.caseid_cashwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet02':
                caseid_stx_app.caseid_cashwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet03':
                caseid_stx_app.caseid_cashwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet04':
                caseid_stx_app.caseid_cashwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet05':
                caseid_stx_app.caseid_cashwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet06':
                caseid_stx_app.caseid_cashwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet07':
                caseid_stx_app.caseid_cashwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet08':
                caseid_stx_app.caseid_cashwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet09':
                caseid_stx_app.caseid_cashwallet09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet10':
                caseid_stx_app.caseid_cashwallet10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'TripPayment01':
                caseid_stx_app.caseid_trippayment01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment02':
                caseid_stx_app.caseid_trippayment02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment03':
                caseid_stx_app.caseid_trippayment03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment04':
                caseid_stx_app.caseid_trippayment04(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'VnPayWallet01':
                caseid_stx_app.caseid_vnpaywallett01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'VnPayWallet02':
                caseid_stx_app.caseid_vnpaywallett02(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Setting01':
                caseid_stx_app.caseid_setting01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting02':
                caseid_stx_app.caseid_setting02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting03':
                caseid_stx_app.caseid_setting03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting04':
                caseid_stx_app.caseid_setting04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting05':
                caseid_stx_app.caseid_setting05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting06':
                caseid_stx_app.caseid_setting06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting07':
                caseid_stx_app.caseid_setting07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting08':
                caseid_stx_app.caseid_setting08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting09':
                caseid_stx_app.caseid_setting09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting10':
                caseid_stx_app.caseid_setting10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Report01':
                caseid_stx_app.caseid_report01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report02':
                caseid_stx_app.caseid_report02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report03':
                caseid_stx_app.caseid_report03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report04':
                caseid_stx_app.caseid_report04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report05':
                caseid_stx_app.caseid_report05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report06':
                caseid_stx_app.caseid_report06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report07':
                caseid_stx_app.caseid_report07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report08':
                caseid_stx_app.caseid_report08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report09':
                caseid_stx_app.caseid_report09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report10':
                caseid_stx_app.caseid_report10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report11':
                caseid_stx_app.caseid_report11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report12':
                caseid_stx_app.caseid_report12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report13':
                caseid_stx_app.caseid_report13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report14':
                caseid_stx_app.caseid_report14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report15':
                caseid_stx_app.caseid_report15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report16':
                caseid_stx_app.caseid_report16(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Help01':
                caseid_stx_app.caseid_help01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help02':
                caseid_stx_app.caseid_help02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help03':
                caseid_stx_app.caseid_help03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help04':
                caseid_stx_app.caseid_help04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help05':
                caseid_stx_app.caseid_help05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Logout01':
                caseid_stx_app.caseid_logout01(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Web01':
                caseid_stx_app.caseid_web01(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web02':
                caseid_stx_app.caseid_web02(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web03':
                caseid_stx_app.caseid_web03(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web04':
                caseid_stx_app.caseid_web04(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web05':
                caseid_stx_app.caseid_web05(self)
        except:
            module_other_stx_app.close_chrome()





def luong_customer(self):
    list_0 = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        print(sheet["I" + rownum].value)
        print(sheet["H" + rownum].value)
        if sheet["I" + rownum].value == "x" and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case = sheet["A" + rownum].value
            list_0.append(case)
        rownum = int(rownum)
    print(list_0)
    print("Tổng số case luồng khách hàng(1): ", len(list_0))
    for case in list_0:
        try:
            if case == 'Login01':
                caseid_stx_app.caseid_login01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login02':
                caseid_stx_app.caseid_login02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login03':
                caseid_stx_app.caseid_login03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login04':
                caseid_stx_app.caseid_login04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login05':
                caseid_stx_app.caseid_login05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login06':
                caseid_stx_app.caseid_login06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login07':
                caseid_stx_app.caseid_login07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login08':
                caseid_stx_app.caseid_login08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login09':
                caseid_stx_app.caseid_login09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login10':
                caseid_stx_app.caseid_login10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage01':
                caseid_stx_app.caseid_homepage01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage02':
                caseid_stx_app.caseid_homepage02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage03':
                caseid_stx_app.caseid_homepage03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage04':
                caseid_stx_app.caseid_homepage04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage05':
                caseid_stx_app.caseid_homepage05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage06':
                caseid_stx_app.caseid_homepage06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage07':
                caseid_stx_app.caseid_homepage07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage08':
                caseid_stx_app.caseid_homepage08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage09':
                caseid_stx_app.caseid_homepage09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage10':
                caseid_stx_app.caseid_homepage10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage11':
                caseid_stx_app.caseid_homepage11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12':
                caseid_stx_app.caseid_homepage12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage12_1':
                caseid_stx_app.caseid_homepage12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12_2':
                caseid_stx_app.caseid_homepage12_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage13':
                caseid_stx_app.caseid_homepage13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage14':
                caseid_stx_app.caseid_homepage14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage15':
                caseid_stx_app.caseid_homepage15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage16':
                caseid_stx_app.caseid_homepage16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage17':
                caseid_stx_app.caseid_homepage17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage18':
                caseid_stx_app.caseid_homepage18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage19':
                caseid_stx_app.caseid_homepage19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage20':
                caseid_stx_app.caseid_homepage20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage21':
                caseid_stx_app.caseid_homepage21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage22':
                caseid_stx_app.caseid_homepage22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage23':
                caseid_stx_app.caseid_homepage23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage24':
                caseid_stx_app.caseid_homepage24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage25':
                caseid_stx_app.caseid_homepage25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage26':
                caseid_stx_app.caseid_homepage26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage27':
                caseid_stx_app.caseid_homepage27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28':
                caseid_stx_app.caseid_homepage28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28_1':
                caseid_stx_app.caseid_homepage28_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage29':
                caseid_stx_app.caseid_homepage29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage30':
                caseid_stx_app.caseid_homepage30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage31':
                caseid_stx_app.caseid_homepage31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage32':
                caseid_stx_app.caseid_homepage32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage33':
                caseid_stx_app.caseid_homepage33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage34':
                caseid_stx_app.caseid_homepage34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage35':
                caseid_stx_app.caseid_homepage35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage36':
                caseid_stx_app.caseid_homepage36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage37':
                caseid_stx_app.caseid_homepage37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage38':
                caseid_stx_app.caseid_homepage38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage39':
                caseid_stx_app.caseid_homepage39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage40':
                caseid_stx_app.caseid_homepage40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage41':
                caseid_stx_app.caseid_homepage41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42':
                caseid_stx_app.caseid_homepage42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42_1':
                caseid_stx_app.caseid_homepage42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43':
                caseid_stx_app.caseid_homepage43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_1':
                caseid_stx_app.caseid_homepage43_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_2':
                caseid_stx_app.caseid_homepage43_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage44':
                caseid_stx_app.caseid_homepage44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage45':
                caseid_stx_app.caseid_homepage45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage46':
                caseid_stx_app.caseid_homepage46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage47':
                caseid_stx_app.caseid_homepage47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage48':
                caseid_stx_app.caseid_homepage48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage49':
                caseid_stx_app.caseid_homepage49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage50':
                caseid_stx_app.caseid_homepage50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage51':
                caseid_stx_app.caseid_homepage51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage52':
                caseid_stx_app.caseid_homepage52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage53':
                caseid_stx_app.caseid_homepage53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage54':
                caseid_stx_app.caseid_homepage54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage55':
                caseid_stx_app.caseid_homepage55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage56':
                caseid_stx_app.caseid_homepage56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage57':
                caseid_stx_app.caseid_homepage57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage58':
                caseid_stx_app.caseid_homepage58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage59':
                caseid_stx_app.caseid_homepage59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage60':
                caseid_stx_app.caseid_homepage60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage61':
                caseid_stx_app.caseid_homepage61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage62':
                caseid_stx_app.caseid_homepage62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage63':
                caseid_stx_app.caseid_homepage63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage64':
                caseid_stx_app.caseid_homepage64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage65':
                caseid_stx_app.caseid_homepage65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage66':
                caseid_stx_app.caseid_homepage66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage67':
                caseid_stx_app.caseid_homepage67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage68':
                caseid_stx_app.caseid_homepage68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage69':
                caseid_stx_app.caseid_homepage69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage70':
                caseid_stx_app.caseid_homepage70(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'FavoriteSpot01':
                caseid_stx_app.caseid_favoritespot01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot02':
                caseid_stx_app.caseid_favoritespot02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot03':
                caseid_stx_app.caseid_favoritespot03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot04':
                caseid_stx_app.caseid_favoritespot04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot05':
                caseid_stx_app.caseid_favoritespot05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot06':
                caseid_stx_app.caseid_favoritespot06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot07':
                caseid_stx_app.caseid_favoritespot07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot08':
                caseid_stx_app.caseid_favoritespot08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot09':
                caseid_stx_app.caseid_favoritespot09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot10':
                caseid_stx_app.caseid_favoritespot10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot11':
                caseid_stx_app.caseid_favoritespot11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot12':
                caseid_stx_app.caseid_favoritespot12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Promotion01':
                caseid_stx_app.caseid_promotion01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion02':
                caseid_stx_app.caseid_promotion02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion03':
                caseid_stx_app.caseid_promotion03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion04':
                caseid_stx_app.caseid_promotion04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion05':
                caseid_stx_app.caseid_promotion05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion06':
                caseid_stx_app.caseid_promotion06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion07':
                caseid_stx_app.caseid_promotion07(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account01':
                caseid_stx_app.caseid_account01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account02':
                caseid_stx_app.caseid_account02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account03':
                caseid_stx_app.caseid_account03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account04':
                caseid_stx_app.caseid_account04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account05':
                caseid_stx_app.caseid_account05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account06':
                caseid_stx_app.caseid_account06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account07':
                caseid_stx_app.caseid_account07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account08':
                caseid_stx_app.caseid_account08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account09':
                caseid_stx_app.caseid_account09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account10':
                caseid_stx_app.caseid_account10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account11':
                caseid_stx_app.caseid_account11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12':
                caseid_stx_app.caseid_account12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_1':
                caseid_stx_app.caseid_account12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_2':
                caseid_stx_app.caseid_account12_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_3':
                caseid_stx_app.caseid_account12_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_4':
                caseid_stx_app.caseid_account12_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_5':
                caseid_stx_app.caseid_account12_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_6':
                caseid_stx_app.caseid_account12_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13':
                caseid_stx_app.caseid_account13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13_1':
                caseid_stx_app.caseid_account13_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account14':
                caseid_stx_app.caseid_account14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account15':
                caseid_stx_app.caseid_account15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account16':
                caseid_stx_app.caseid_account16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account17':
                caseid_stx_app.caseid_account17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account18':
                caseid_stx_app.caseid_account18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account19':
                caseid_stx_app.caseid_account19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account20':
                caseid_stx_app.caseid_account20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account21':
                caseid_stx_app.caseid_account21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account22':
                caseid_stx_app.caseid_account22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account23':
                caseid_stx_app.caseid_account23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account24':
                caseid_stx_app.caseid_account24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account25':
                caseid_stx_app.caseid_account25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account26':
                caseid_stx_app.caseid_account26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account27':
                caseid_stx_app.caseid_account27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account28':
                caseid_stx_app.caseid_account28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account29':
                caseid_stx_app.caseid_account29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account30':
                caseid_stx_app.caseid_account30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account31':
                caseid_stx_app.caseid_account31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account32':
                caseid_stx_app.caseid_account32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account33':
                caseid_stx_app.caseid_account33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account34':
                caseid_stx_app.caseid_account34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account35':
                caseid_stx_app.caseid_account35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account36':
                caseid_stx_app.caseid_account36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account37':
                caseid_stx_app.caseid_account37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account38':
                caseid_stx_app.caseid_account38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account39':
                caseid_stx_app.caseid_account39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40':
                caseid_stx_app.caseid_account40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40_1':
                caseid_stx_app.caseid_account40_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account41':
                caseid_stx_app.caseid_account41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account42':
                caseid_stx_app.caseid_account42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account43':
                caseid_stx_app.caseid_account43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account44':
                caseid_stx_app.caseid_account44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account45':
                caseid_stx_app.caseid_account45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account46':
                caseid_stx_app.caseid_account46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account47':
                caseid_stx_app.caseid_account47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account48':
                caseid_stx_app.caseid_account48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account49':
                caseid_stx_app.caseid_account49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50':
                caseid_stx_app.caseid_account50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50_1':
                caseid_stx_app.caseid_account50_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51':
                caseid_stx_app.caseid_account51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_1':
                caseid_stx_app.caseid_account51_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_2':
                caseid_stx_app.caseid_account51_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account52':
                caseid_stx_app.caseid_account52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account53':
                caseid_stx_app.caseid_account53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account54':
                caseid_stx_app.caseid_account54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account55':
                caseid_stx_app.caseid_account55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account56':
                caseid_stx_app.caseid_account56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account57':
                caseid_stx_app.caseid_account57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account58':
                caseid_stx_app.caseid_account58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account59':
                caseid_stx_app.caseid_account59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account60':
                caseid_stx_app.caseid_account60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account61':
                caseid_stx_app.caseid_account61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account62':
                caseid_stx_app.caseid_account62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account63':
                caseid_stx_app.caseid_account63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account64':
                caseid_stx_app.caseid_account64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account65':
                caseid_stx_app.caseid_account65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account66':
                caseid_stx_app.caseid_account66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account67':
                caseid_stx_app.caseid_account67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account68':
                caseid_stx_app.caseid_account68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account69':
                caseid_stx_app.caseid_account69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account70':
                caseid_stx_app.caseid_account70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account71':
                caseid_stx_app.caseid_account71(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'App01':
                caseid_stx_app.caseid_app01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App02':
                caseid_stx_app.caseid_app02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App03':
                caseid_stx_app.caseid_app03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App04':
                caseid_stx_app.caseid_app04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App05':
                caseid_stx_app.caseid_app05(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home01':
                caseid_stx_app.caseid_home01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home02':
                caseid_stx_app.caseid_home02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home03':
                caseid_stx_app.caseid_home03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home04':
                caseid_stx_app.caseid_home04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home05':
                caseid_stx_app.caseid_home05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home06':
                caseid_stx_app.caseid_home06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home07':
                caseid_stx_app.caseid_home07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home08':
                caseid_stx_app.caseid_home08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09':
                caseid_stx_app.caseid_home09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home09_1':
                caseid_stx_app.caseid_home09_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09_2':
                caseid_stx_app.caseid_home09_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home10':
                caseid_stx_app.caseid_home10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home11':
                caseid_stx_app.caseid_home11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home12':
                caseid_stx_app.caseid_home12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home13':
                caseid_stx_app.caseid_home13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home14':
                caseid_stx_app.caseid_home14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home15':
                caseid_stx_app.caseid_home15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home16':
                caseid_stx_app.caseid_home16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home17':
                caseid_stx_app.caseid_home17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home18':
                caseid_stx_app.caseid_home18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home19':
                caseid_stx_app.caseid_home19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home20':
                caseid_stx_app.caseid_home20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home21':
                caseid_stx_app.caseid_home21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home22':
                caseid_stx_app.caseid_home22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home23':
                caseid_stx_app.caseid_home23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home24':
                caseid_stx_app.caseid_home24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home25':
                caseid_stx_app.caseid_home25(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home25_1':
                caseid_stx_app.caseid_home25_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home26':
                caseid_stx_app.caseid_home26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home27':
                caseid_stx_app.caseid_home27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home28':
                caseid_stx_app.caseid_home28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home29':
                caseid_stx_app.caseid_home29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30':
                caseid_stx_app.caseid_home30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30_1':
                caseid_stx_app.caseid_home30_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home31':
                caseid_stx_app.caseid_home31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home32':
                caseid_stx_app.caseid_home32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home33':
                caseid_stx_app.caseid_home33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home34':
                caseid_stx_app.caseid_home34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home35':
                caseid_stx_app.caseid_home35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home36':
                caseid_stx_app.caseid_home36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home37':
                caseid_stx_app.caseid_home37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home38':
                caseid_stx_app.caseid_home38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home39':
                caseid_stx_app.caseid_home39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home40':
                caseid_stx_app.caseid_home40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home41':
                caseid_stx_app.caseid_home41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42':
                caseid_stx_app.caseid_home42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_1':
                caseid_stx_app.caseid_home42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_2':
                caseid_stx_app.caseid_home42_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_3':
                caseid_stx_app.caseid_home42_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_4':
                caseid_stx_app.caseid_home42_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_5':
                caseid_stx_app.caseid_home42_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_6':
                caseid_stx_app.caseid_home42_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_7':
                caseid_stx_app.caseid_home42_7(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_8':
                caseid_stx_app.caseid_home42_8(self)
        except:
            module_other_stx_app.close_app()



        try:
            if case == 'Home43':
                caseid_stx_app.caseid_home43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home43_1':
                caseid_stx_app.caseid_home43_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home44':
                caseid_stx_app.caseid_home44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home45':
                caseid_stx_app.caseid_home45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home46':
                caseid_stx_app.caseid_home46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47':
                caseid_stx_app.caseid_home47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_1':
                caseid_stx_app.caseid_home47_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_2':
                caseid_stx_app.caseid_home47_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_3':
                caseid_stx_app.caseid_home47_3(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home48':
                caseid_stx_app.caseid_home48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home49':
                caseid_stx_app.caseid_home49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home50':
                caseid_stx_app.caseid_home50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home51':
                caseid_stx_app.caseid_home51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home52':
                caseid_stx_app.caseid_home52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home53':
                caseid_stx_app.caseid_home53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home54':
                caseid_stx_app.caseid_home54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home55':
                caseid_stx_app.caseid_home55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home56':
                caseid_stx_app.caseid_home56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home57':
                caseid_stx_app.caseid_home57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home58':
                caseid_stx_app.caseid_home58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home59':
                caseid_stx_app.caseid_home59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home60':
                caseid_stx_app.caseid_home60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home61':
                caseid_stx_app.caseid_home61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home62':
                caseid_stx_app.caseid_home62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home63':
                caseid_stx_app.caseid_home63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home64':
                caseid_stx_app.caseid_home64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home65':
                caseid_stx_app.caseid_home65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home66':
                caseid_stx_app.caseid_home66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home67':
                caseid_stx_app.caseid_home67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home68':
                caseid_stx_app.caseid_home68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home69':
                caseid_stx_app.caseid_home69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home70':
                caseid_stx_app.caseid_home70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home71':
                caseid_stx_app.caseid_home71(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home72':
                caseid_stx_app.caseid_home72(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home73':
                caseid_stx_app.caseid_home73(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home74':
                caseid_stx_app.caseid_home74(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home75':
                caseid_stx_app.caseid_home75(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home76':
                caseid_stx_app.caseid_home76(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home77':
                caseid_stx_app.caseid_home77(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home78':
                caseid_stx_app.caseid_home78(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home79':
                caseid_stx_app.caseid_home79(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home80':
                caseid_stx_app.caseid_home80(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home81':
                caseid_stx_app.caseid_home81(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home82':
                caseid_stx_app.caseid_home82(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Notification01':
                caseid_stx_app.caseid_notification01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification02':
                caseid_stx_app.caseid_notification02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification03':
                caseid_stx_app.caseid_notification03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification04':
                caseid_stx_app.caseid_notification04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification05':
                caseid_stx_app.caseid_notification05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification06':
                caseid_stx_app.caseid_notification06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification07':
                caseid_stx_app.caseid_notification07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification08':
                caseid_stx_app.caseid_notification08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification09':
                caseid_stx_app.caseid_notification09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification10':
                caseid_stx_app.caseid_notification10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History01':
                caseid_stx_app.caseid_history01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02':
                caseid_stx_app.caseid_history02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02_1':
                caseid_stx_app.caseid_history02_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History03':
                caseid_stx_app.caseid_history03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History04':
                caseid_stx_app.caseid_history04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History05':
                caseid_stx_app.caseid_history05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History06':
                caseid_stx_app.caseid_history06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History07':
                caseid_stx_app.caseid_history07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History08':
                caseid_stx_app.caseid_history08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History09':
                caseid_stx_app.caseid_history09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History10':
                caseid_stx_app.caseid_history10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History11':
                caseid_stx_app.caseid_history11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History12':
                caseid_stx_app.caseid_history12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13':
                caseid_stx_app.caseid_history13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_1':
                caseid_stx_app.caseid_history13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_2':
                caseid_stx_app.caseid_history13_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_3':
                caseid_stx_app.caseid_history13_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_4':
                caseid_stx_app.caseid_history13_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_5':
                caseid_stx_app.caseid_history13_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History14':
                caseid_stx_app.caseid_history14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History15':
                caseid_stx_app.caseid_history15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History16':
                caseid_stx_app.caseid_history16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History17':
                caseid_stx_app.caseid_history17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History18':
                caseid_stx_app.caseid_history18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19':
                caseid_stx_app.caseid_history19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19_1':
                caseid_stx_app.caseid_history19_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History20':
                caseid_stx_app.caseid_history20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History21':
                caseid_stx_app.caseid_history21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History22':
                caseid_stx_app.caseid_history22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History23':
                caseid_stx_app.caseid_history23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History24':
                caseid_stx_app.caseid_history24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History25':
                caseid_stx_app.caseid_history25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History26':
                caseid_stx_app.caseid_history26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History27':
                caseid_stx_app.caseid_history27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History28':
                caseid_stx_app.caseid_history28(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'AccountWallet01':
                caseid_stx_app.caseid_accountwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet02':
                caseid_stx_app.caseid_accountwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet03':
                caseid_stx_app.caseid_accountwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet04':
                caseid_stx_app.caseid_accountwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet05':
                caseid_stx_app.caseid_accountwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet06':
                caseid_stx_app.caseid_accountwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet07':
                caseid_stx_app.caseid_accountwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet08':
                caseid_stx_app.caseid_accountwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet09':
                caseid_stx_app.caseid_accountwallet09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'CashWallet01':
                caseid_stx_app.caseid_cashwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet02':
                caseid_stx_app.caseid_cashwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet03':
                caseid_stx_app.caseid_cashwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet04':
                caseid_stx_app.caseid_cashwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet05':
                caseid_stx_app.caseid_cashwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet06':
                caseid_stx_app.caseid_cashwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet07':
                caseid_stx_app.caseid_cashwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet08':
                caseid_stx_app.caseid_cashwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet09':
                caseid_stx_app.caseid_cashwallet09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet10':
                caseid_stx_app.caseid_cashwallet10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'TripPayment01':
                caseid_stx_app.caseid_trippayment01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment02':
                caseid_stx_app.caseid_trippayment02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment03':
                caseid_stx_app.caseid_trippayment03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment04':
                caseid_stx_app.caseid_trippayment04(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'VnPayWallet01':
                caseid_stx_app.caseid_vnpaywallett01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'VnPayWallet02':
                caseid_stx_app.caseid_vnpaywallett02(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Setting01':
                caseid_stx_app.caseid_setting01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting02':
                caseid_stx_app.caseid_setting02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting03':
                caseid_stx_app.caseid_setting03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting04':
                caseid_stx_app.caseid_setting04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting05':
                caseid_stx_app.caseid_setting05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting06':
                caseid_stx_app.caseid_setting06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting07':
                caseid_stx_app.caseid_setting07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting08':
                caseid_stx_app.caseid_setting08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting09':
                caseid_stx_app.caseid_setting09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting10':
                caseid_stx_app.caseid_setting10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Report01':
                caseid_stx_app.caseid_report01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report02':
                caseid_stx_app.caseid_report02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report03':
                caseid_stx_app.caseid_report03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report04':
                caseid_stx_app.caseid_report04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report05':
                caseid_stx_app.caseid_report05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report06':
                caseid_stx_app.caseid_report06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report07':
                caseid_stx_app.caseid_report07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report08':
                caseid_stx_app.caseid_report08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report09':
                caseid_stx_app.caseid_report09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report10':
                caseid_stx_app.caseid_report10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report11':
                caseid_stx_app.caseid_report11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report12':
                caseid_stx_app.caseid_report12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report13':
                caseid_stx_app.caseid_report13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report14':
                caseid_stx_app.caseid_report14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report15':
                caseid_stx_app.caseid_report15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report16':
                caseid_stx_app.caseid_report16(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Help01':
                caseid_stx_app.caseid_help01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help02':
                caseid_stx_app.caseid_help02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help03':
                caseid_stx_app.caseid_help03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help04':
                caseid_stx_app.caseid_help04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help05':
                caseid_stx_app.caseid_help05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Logout01':
                caseid_stx_app.caseid_logout01(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Web01':
                caseid_stx_app.caseid_web01(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web02':
                caseid_stx_app.caseid_web02(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web03':
                caseid_stx_app.caseid_web03(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web04':
                caseid_stx_app.caseid_web04(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web05':
                caseid_stx_app.caseid_web05(self)
        except:
            module_other_stx_app.close_chrome()




def luong_drive(self):
    list_0 = []
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 600):
        rownum += 1
        rownum = str(rownum)
        print(sheet["J" + rownum].value)
        print(sheet["H" + rownum].value)
        if sheet["J" + rownum].value == "x" and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case = sheet["A" + rownum].value
            list_0.append(case)
        rownum = int(rownum)
    print(list_0)
    print("Tổng số case luồng tài xế(2): ", len(list_0))
    for case in list_0:
        try:
            if case == 'Login01':
                caseid_stx_app.caseid_login01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login02':
                caseid_stx_app.caseid_login02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login03':
                caseid_stx_app.caseid_login03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login04':
                caseid_stx_app.caseid_login04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login05':
                caseid_stx_app.caseid_login05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login06':
                caseid_stx_app.caseid_login06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login07':
                caseid_stx_app.caseid_login07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login08':
                caseid_stx_app.caseid_login08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login09':
                caseid_stx_app.caseid_login09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Login10':
                caseid_stx_app.caseid_login10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage01':
                caseid_stx_app.caseid_homepage01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage02':
                caseid_stx_app.caseid_homepage02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage03':
                caseid_stx_app.caseid_homepage03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage04':
                caseid_stx_app.caseid_homepage04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage05':
                caseid_stx_app.caseid_homepage05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage06':
                caseid_stx_app.caseid_homepage06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage07':
                caseid_stx_app.caseid_homepage07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage08':
                caseid_stx_app.caseid_homepage08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage09':
                caseid_stx_app.caseid_homepage09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage10':
                caseid_stx_app.caseid_homepage10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage11':
                caseid_stx_app.caseid_homepage11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12':
                caseid_stx_app.caseid_homepage12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'HomePage12_1':
                caseid_stx_app.caseid_homepage12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage12_2':
                caseid_stx_app.caseid_homepage12_2(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'HomePage13':
                caseid_stx_app.caseid_homepage13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage14':
                caseid_stx_app.caseid_homepage14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage15':
                caseid_stx_app.caseid_homepage15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage16':
                caseid_stx_app.caseid_homepage16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage17':
                caseid_stx_app.caseid_homepage17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage18':
                caseid_stx_app.caseid_homepage18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage19':
                caseid_stx_app.caseid_homepage19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage20':
                caseid_stx_app.caseid_homepage20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage21':
                caseid_stx_app.caseid_homepage21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage22':
                caseid_stx_app.caseid_homepage22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage23':
                caseid_stx_app.caseid_homepage23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage24':
                caseid_stx_app.caseid_homepage24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage25':
                caseid_stx_app.caseid_homepage25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage26':
                caseid_stx_app.caseid_homepage26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage27':
                caseid_stx_app.caseid_homepage27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28':
                caseid_stx_app.caseid_homepage28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage28_1':
                caseid_stx_app.caseid_homepage28_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage29':
                caseid_stx_app.caseid_homepage29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage30':
                caseid_stx_app.caseid_homepage30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage31':
                caseid_stx_app.caseid_homepage31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage32':
                caseid_stx_app.caseid_homepage32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage33':
                caseid_stx_app.caseid_homepage33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage34':
                caseid_stx_app.caseid_homepage34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage35':
                caseid_stx_app.caseid_homepage35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage36':
                caseid_stx_app.caseid_homepage36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage37':
                caseid_stx_app.caseid_homepage37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage38':
                caseid_stx_app.caseid_homepage38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage39':
                caseid_stx_app.caseid_homepage39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage40':
                caseid_stx_app.caseid_homepage40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage41':
                caseid_stx_app.caseid_homepage41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42':
                caseid_stx_app.caseid_homepage42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage42_1':
                caseid_stx_app.caseid_homepage42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43':
                caseid_stx_app.caseid_homepage43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_1':
                caseid_stx_app.caseid_homepage43_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage43_2':
                caseid_stx_app.caseid_homepage43_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage44':
                caseid_stx_app.caseid_homepage44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage45':
                caseid_stx_app.caseid_homepage45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage46':
                caseid_stx_app.caseid_homepage46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage47':
                caseid_stx_app.caseid_homepage47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage48':
                caseid_stx_app.caseid_homepage48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage49':
                caseid_stx_app.caseid_homepage49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage50':
                caseid_stx_app.caseid_homepage50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage51':
                caseid_stx_app.caseid_homepage51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage52':
                caseid_stx_app.caseid_homepage52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage53':
                caseid_stx_app.caseid_homepage53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage54':
                caseid_stx_app.caseid_homepage54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage55':
                caseid_stx_app.caseid_homepage55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage56':
                caseid_stx_app.caseid_homepage56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage57':
                caseid_stx_app.caseid_homepage57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage58':
                caseid_stx_app.caseid_homepage58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage59':
                caseid_stx_app.caseid_homepage59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage60':
                caseid_stx_app.caseid_homepage60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage61':
                caseid_stx_app.caseid_homepage61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage62':
                caseid_stx_app.caseid_homepage62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage63':
                caseid_stx_app.caseid_homepage63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage64':
                caseid_stx_app.caseid_homepage64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage65':
                caseid_stx_app.caseid_homepage65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage66':
                caseid_stx_app.caseid_homepage66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage67':
                caseid_stx_app.caseid_homepage67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage68':
                caseid_stx_app.caseid_homepage68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage69':
                caseid_stx_app.caseid_homepage69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'HomePage70':
                caseid_stx_app.caseid_homepage70(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'FavoriteSpot01':
                caseid_stx_app.caseid_favoritespot01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot02':
                caseid_stx_app.caseid_favoritespot02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot03':
                caseid_stx_app.caseid_favoritespot03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot04':
                caseid_stx_app.caseid_favoritespot04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot05':
                caseid_stx_app.caseid_favoritespot05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot06':
                caseid_stx_app.caseid_favoritespot06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot07':
                caseid_stx_app.caseid_favoritespot07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot08':
                caseid_stx_app.caseid_favoritespot08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot09':
                caseid_stx_app.caseid_favoritespot09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot10':
                caseid_stx_app.caseid_favoritespot10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot11':
                caseid_stx_app.caseid_favoritespot11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'FavoriteSpot12':
                caseid_stx_app.caseid_favoritespot12(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Promotion01':
                caseid_stx_app.caseid_promotion01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion02':
                caseid_stx_app.caseid_promotion02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion03':
                caseid_stx_app.caseid_promotion03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion04':
                caseid_stx_app.caseid_promotion04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion05':
                caseid_stx_app.caseid_promotion05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion06':
                caseid_stx_app.caseid_promotion06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Promotion07':
                caseid_stx_app.caseid_promotion07(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account01':
                caseid_stx_app.caseid_account01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account02':
                caseid_stx_app.caseid_account02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account03':
                caseid_stx_app.caseid_account03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account04':
                caseid_stx_app.caseid_account04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account05':
                caseid_stx_app.caseid_account05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account06':
                caseid_stx_app.caseid_account06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account07':
                caseid_stx_app.caseid_account07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account08':
                caseid_stx_app.caseid_account08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account09':
                caseid_stx_app.caseid_account09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account10':
                caseid_stx_app.caseid_account10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account11':
                caseid_stx_app.caseid_account11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12':
                caseid_stx_app.caseid_account12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_1':
                caseid_stx_app.caseid_account12_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_2':
                caseid_stx_app.caseid_account12_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_3':
                caseid_stx_app.caseid_account12_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_4':
                caseid_stx_app.caseid_account12_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_5':
                caseid_stx_app.caseid_account12_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account12_6':
                caseid_stx_app.caseid_account12_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13':
                caseid_stx_app.caseid_account13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account13_1':
                caseid_stx_app.caseid_account13_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Account14':
                caseid_stx_app.caseid_account14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account15':
                caseid_stx_app.caseid_account15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account16':
                caseid_stx_app.caseid_account16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account17':
                caseid_stx_app.caseid_account17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account18':
                caseid_stx_app.caseid_account18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account19':
                caseid_stx_app.caseid_account19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account20':
                caseid_stx_app.caseid_account20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account21':
                caseid_stx_app.caseid_account21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account22':
                caseid_stx_app.caseid_account22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account23':
                caseid_stx_app.caseid_account23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account24':
                caseid_stx_app.caseid_account24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account25':
                caseid_stx_app.caseid_account25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account26':
                caseid_stx_app.caseid_account26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account27':
                caseid_stx_app.caseid_account27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account28':
                caseid_stx_app.caseid_account28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account29':
                caseid_stx_app.caseid_account29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account30':
                caseid_stx_app.caseid_account30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account31':
                caseid_stx_app.caseid_account31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account32':
                caseid_stx_app.caseid_account32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account33':
                caseid_stx_app.caseid_account33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account34':
                caseid_stx_app.caseid_account34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account35':
                caseid_stx_app.caseid_account35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account36':
                caseid_stx_app.caseid_account36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account37':
                caseid_stx_app.caseid_account37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account38':
                caseid_stx_app.caseid_account38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account39':
                caseid_stx_app.caseid_account39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40':
                caseid_stx_app.caseid_account40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account40_1':
                caseid_stx_app.caseid_account40_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account41':
                caseid_stx_app.caseid_account41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account42':
                caseid_stx_app.caseid_account42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account43':
                caseid_stx_app.caseid_account43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account44':
                caseid_stx_app.caseid_account44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account45':
                caseid_stx_app.caseid_account45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account46':
                caseid_stx_app.caseid_account46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account47':
                caseid_stx_app.caseid_account47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account48':
                caseid_stx_app.caseid_account48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account49':
                caseid_stx_app.caseid_account49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50':
                caseid_stx_app.caseid_account50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account50_1':
                caseid_stx_app.caseid_account50_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51':
                caseid_stx_app.caseid_account51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_1':
                caseid_stx_app.caseid_account51_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account51_2':
                caseid_stx_app.caseid_account51_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account52':
                caseid_stx_app.caseid_account52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account53':
                caseid_stx_app.caseid_account53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account54':
                caseid_stx_app.caseid_account54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account55':
                caseid_stx_app.caseid_account55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account56':
                caseid_stx_app.caseid_account56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account57':
                caseid_stx_app.caseid_account57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account58':
                caseid_stx_app.caseid_account58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account59':
                caseid_stx_app.caseid_account59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account60':
                caseid_stx_app.caseid_account60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account61':
                caseid_stx_app.caseid_account61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account62':
                caseid_stx_app.caseid_account62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account63':
                caseid_stx_app.caseid_account63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account64':
                caseid_stx_app.caseid_account64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account65':
                caseid_stx_app.caseid_account65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account66':
                caseid_stx_app.caseid_account66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account67':
                caseid_stx_app.caseid_account67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account68':
                caseid_stx_app.caseid_account68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account69':
                caseid_stx_app.caseid_account69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account70':
                caseid_stx_app.caseid_account70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Account71':
                caseid_stx_app.caseid_account71(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'App01':
                caseid_stx_app.caseid_app01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App02':
                caseid_stx_app.caseid_app02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App03':
                caseid_stx_app.caseid_app03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App04':
                caseid_stx_app.caseid_app04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'App05':
                caseid_stx_app.caseid_app05(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home01':
                caseid_stx_app.caseid_home01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home02':
                caseid_stx_app.caseid_home02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home03':
                caseid_stx_app.caseid_home03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home04':
                caseid_stx_app.caseid_home04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home05':
                caseid_stx_app.caseid_home05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home06':
                caseid_stx_app.caseid_home06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home07':
                caseid_stx_app.caseid_home07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home08':
                caseid_stx_app.caseid_home08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09':
                caseid_stx_app.caseid_home09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home09_1':
                caseid_stx_app.caseid_home09_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home09_2':
                caseid_stx_app.caseid_home09_2(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home10':
                caseid_stx_app.caseid_home10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home11':
                caseid_stx_app.caseid_home11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home12':
                caseid_stx_app.caseid_home12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home13':
                caseid_stx_app.caseid_home13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home14':
                caseid_stx_app.caseid_home14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home15':
                caseid_stx_app.caseid_home15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home16':
                caseid_stx_app.caseid_home16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home17':
                caseid_stx_app.caseid_home17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home18':
                caseid_stx_app.caseid_home18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home19':
                caseid_stx_app.caseid_home19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home20':
                caseid_stx_app.caseid_home20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home21':
                caseid_stx_app.caseid_home21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home22':
                caseid_stx_app.caseid_home22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home23':
                caseid_stx_app.caseid_home23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home24':
                caseid_stx_app.caseid_home24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home25':
                caseid_stx_app.caseid_home25(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home25_1':
                caseid_stx_app.caseid_home25_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home26':
                caseid_stx_app.caseid_home26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home27':
                caseid_stx_app.caseid_home27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home28':
                caseid_stx_app.caseid_home28(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home29':
                caseid_stx_app.caseid_home29(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30':
                caseid_stx_app.caseid_home30(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home30_1':
                caseid_stx_app.caseid_home30_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home31':
                caseid_stx_app.caseid_home31(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home32':
                caseid_stx_app.caseid_home32(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home33':
                caseid_stx_app.caseid_home33(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home34':
                caseid_stx_app.caseid_home34(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home35':
                caseid_stx_app.caseid_home35(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home36':
                caseid_stx_app.caseid_home36(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home37':
                caseid_stx_app.caseid_home37(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home38':
                caseid_stx_app.caseid_home38(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home39':
                caseid_stx_app.caseid_home39(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home40':
                caseid_stx_app.caseid_home40(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home41':
                caseid_stx_app.caseid_home41(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42':
                caseid_stx_app.caseid_home42(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_1':
                caseid_stx_app.caseid_home42_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_2':
                caseid_stx_app.caseid_home42_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_3':
                caseid_stx_app.caseid_home42_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_4':
                caseid_stx_app.caseid_home42_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_5':
                caseid_stx_app.caseid_home42_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_6':
                caseid_stx_app.caseid_home42_6(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_7':
                caseid_stx_app.caseid_home42_7(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home42_8':
                caseid_stx_app.caseid_home42_8(self)
        except:
            module_other_stx_app.close_app()




        try:
            if case == 'Home43':
                caseid_stx_app.caseid_home43(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home43_1':
                caseid_stx_app.caseid_home43_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home44':
                caseid_stx_app.caseid_home44(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home45':
                caseid_stx_app.caseid_home45(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home46':
                caseid_stx_app.caseid_home46(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47':
                caseid_stx_app.caseid_home47(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_1':
                caseid_stx_app.caseid_home47_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_2':
                caseid_stx_app.caseid_home47_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home47_3':
                caseid_stx_app.caseid_home47_3(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Home48':
                caseid_stx_app.caseid_home48(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home49':
                caseid_stx_app.caseid_home49(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home50':
                caseid_stx_app.caseid_home50(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home51':
                caseid_stx_app.caseid_home51(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home52':
                caseid_stx_app.caseid_home52(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home53':
                caseid_stx_app.caseid_home53(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home54':
                caseid_stx_app.caseid_home54(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home55':
                caseid_stx_app.caseid_home55(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home56':
                caseid_stx_app.caseid_home56(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home57':
                caseid_stx_app.caseid_home57(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home58':
                caseid_stx_app.caseid_home58(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home59':
                caseid_stx_app.caseid_home59(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home60':
                caseid_stx_app.caseid_home60(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home61':
                caseid_stx_app.caseid_home61(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home62':
                caseid_stx_app.caseid_home62(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home63':
                caseid_stx_app.caseid_home63(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home64':
                caseid_stx_app.caseid_home64(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home65':
                caseid_stx_app.caseid_home65(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home66':
                caseid_stx_app.caseid_home66(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home67':
                caseid_stx_app.caseid_home67(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home68':
                caseid_stx_app.caseid_home68(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home69':
                caseid_stx_app.caseid_home69(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home70':
                caseid_stx_app.caseid_home70(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home71':
                caseid_stx_app.caseid_home71(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home72':
                caseid_stx_app.caseid_home72(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home73':
                caseid_stx_app.caseid_home73(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home74':
                caseid_stx_app.caseid_home74(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home75':
                caseid_stx_app.caseid_home75(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home76':
                caseid_stx_app.caseid_home76(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home77':
                caseid_stx_app.caseid_home77(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home78':
                caseid_stx_app.caseid_home78(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home79':
                caseid_stx_app.caseid_home79(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home80':
                caseid_stx_app.caseid_home80(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home81':
                caseid_stx_app.caseid_home81(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Home82':
                caseid_stx_app.caseid_home82(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Notification01':
                caseid_stx_app.caseid_notification01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification02':
                caseid_stx_app.caseid_notification02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification03':
                caseid_stx_app.caseid_notification03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification04':
                caseid_stx_app.caseid_notification04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification05':
                caseid_stx_app.caseid_notification05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification06':
                caseid_stx_app.caseid_notification06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification07':
                caseid_stx_app.caseid_notification07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification08':
                caseid_stx_app.caseid_notification08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification09':
                caseid_stx_app.caseid_notification09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Notification10':
                caseid_stx_app.caseid_notification10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History01':
                caseid_stx_app.caseid_history01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02':
                caseid_stx_app.caseid_history02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History02_1':
                caseid_stx_app.caseid_history02_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History03':
                caseid_stx_app.caseid_history03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History04':
                caseid_stx_app.caseid_history04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History05':
                caseid_stx_app.caseid_history05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History06':
                caseid_stx_app.caseid_history06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History07':
                caseid_stx_app.caseid_history07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History08':
                caseid_stx_app.caseid_history08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History09':
                caseid_stx_app.caseid_history09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History10':
                caseid_stx_app.caseid_history10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History11':
                caseid_stx_app.caseid_history11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History12':
                caseid_stx_app.caseid_history12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13':
                caseid_stx_app.caseid_history13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_1':
                caseid_stx_app.caseid_history13_1(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_2':
                caseid_stx_app.caseid_history13_2(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_3':
                caseid_stx_app.caseid_history13_3(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_4':
                caseid_stx_app.caseid_history13_4(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History13_5':
                caseid_stx_app.caseid_history13_5(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History14':
                caseid_stx_app.caseid_history14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History15':
                caseid_stx_app.caseid_history15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History16':
                caseid_stx_app.caseid_history16(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History17':
                caseid_stx_app.caseid_history17(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History18':
                caseid_stx_app.caseid_history18(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19':
                caseid_stx_app.caseid_history19(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History19_1':
                caseid_stx_app.caseid_history19_1(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'History20':
                caseid_stx_app.caseid_history20(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History21':
                caseid_stx_app.caseid_history21(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History22':
                caseid_stx_app.caseid_history22(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History23':
                caseid_stx_app.caseid_history23(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History24':
                caseid_stx_app.caseid_history24(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History25':
                caseid_stx_app.caseid_history25(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History26':
                caseid_stx_app.caseid_history26(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History27':
                caseid_stx_app.caseid_history27(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'History28':
                caseid_stx_app.caseid_history28(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'AccountWallet01':
                caseid_stx_app.caseid_accountwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet02':
                caseid_stx_app.caseid_accountwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet03':
                caseid_stx_app.caseid_accountwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet04':
                caseid_stx_app.caseid_accountwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet05':
                caseid_stx_app.caseid_accountwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet06':
                caseid_stx_app.caseid_accountwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet07':
                caseid_stx_app.caseid_accountwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet08':
                caseid_stx_app.caseid_accountwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'AccountWallet09':
                caseid_stx_app.caseid_accountwallet09(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'CashWallet01':
                caseid_stx_app.caseid_cashwallet01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet02':
                caseid_stx_app.caseid_cashwallet02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet03':
                caseid_stx_app.caseid_cashwallet03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet04':
                caseid_stx_app.caseid_cashwallet04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet05':
                caseid_stx_app.caseid_cashwallet05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet06':
                caseid_stx_app.caseid_cashwallet06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet07':
                caseid_stx_app.caseid_cashwallet07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet08':
                caseid_stx_app.caseid_cashwallet08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet09':
                caseid_stx_app.caseid_cashwallet09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'CashWallet10':
                caseid_stx_app.caseid_cashwallet10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'TripPayment01':
                caseid_stx_app.caseid_trippayment01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment02':
                caseid_stx_app.caseid_trippayment02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment03':
                caseid_stx_app.caseid_trippayment03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'TripPayment04':
                caseid_stx_app.caseid_trippayment04(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'VnPayWallet01':
                caseid_stx_app.caseid_vnpaywallett01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'VnPayWallet02':
                caseid_stx_app.caseid_vnpaywallett02(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Setting01':
                caseid_stx_app.caseid_setting01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting02':
                caseid_stx_app.caseid_setting02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting03':
                caseid_stx_app.caseid_setting03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting04':
                caseid_stx_app.caseid_setting04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting05':
                caseid_stx_app.caseid_setting05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting06':
                caseid_stx_app.caseid_setting06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting07':
                caseid_stx_app.caseid_setting07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting08':
                caseid_stx_app.caseid_setting08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting09':
                caseid_stx_app.caseid_setting09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Setting10':
                caseid_stx_app.caseid_setting10(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Report01':
                caseid_stx_app.caseid_report01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report02':
                caseid_stx_app.caseid_report02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report03':
                caseid_stx_app.caseid_report03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report04':
                caseid_stx_app.caseid_report04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report05':
                caseid_stx_app.caseid_report05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report06':
                caseid_stx_app.caseid_report06(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report07':
                caseid_stx_app.caseid_report07(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report08':
                caseid_stx_app.caseid_report08(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report09':
                caseid_stx_app.caseid_report09(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report10':
                caseid_stx_app.caseid_report10(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report11':
                caseid_stx_app.caseid_report11(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report12':
                caseid_stx_app.caseid_report12(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report13':
                caseid_stx_app.caseid_report13(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report14':
                caseid_stx_app.caseid_report14(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report15':
                caseid_stx_app.caseid_report15(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Report16':
                caseid_stx_app.caseid_report16(self)
        except:
            module_other_stx_app.close_app()


        try:
            if case == 'Help01':
                caseid_stx_app.caseid_help01(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help02':
                caseid_stx_app.caseid_help02(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help03':
                caseid_stx_app.caseid_help03(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help04':
                caseid_stx_app.caseid_help04(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Help05':
                caseid_stx_app.caseid_help05(self)
        except:
            module_other_stx_app.close_app()
        try:
            if case == 'Logout01':
                caseid_stx_app.caseid_logout01(self)
        except:
            module_other_stx_app.close_app()

        try:
            if case == 'Web01':
                caseid_stx_app.caseid_web01(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web02':
                caseid_stx_app.caseid_web02(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web03':
                caseid_stx_app.caseid_web03(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web04':
                caseid_stx_app.caseid_web04(self)
        except:
            module_other_stx_app.close_chrome()
        try:
            if case == 'Web05':
                caseid_stx_app.caseid_web05(self)
        except:
            module_other_stx_app.close_chrome()



