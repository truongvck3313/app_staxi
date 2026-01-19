import time
import openpyxl
import os
from selenium.common.exceptions import NoSuchElementException
import subprocess
import var_stx_app
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import logging
from retry import retry
import module_stx_app
import homepage_g7
import homepage_driver
from pynput.keyboard import Key, Controller
import pygetwindow as gw
import pyautogui
from pynput.keyboard import Controller, Key
import time
import re
import json
import requests
from requests.auth import HTTPBasicAuth
from PIL import Image, ImageEnhance
import pytesseract
import mouse
import time
import var_stx_app
import action
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
import cv2
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

keyboard = Controller()

def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename="stx_app.log",
                                                     encoding='utf-8', mode='w')],  #w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)





import cv2
import numpy as np
from PIL import ImageDraw


def check_icon(template_path, region=None, threshold=0.85, click=False, show=False, retries=3, delay=3):
    """
    Tìm icon trên màn hình bằng pyautogui
    - template_path: đường dẫn ảnh icon cần tìm
    - region: (x, y, w, h) khu vực cần tìm (None = toàn màn hình)
    - threshold: độ tin cậy (0.0 ~ 1.0)
    - click: True thì click vào icon nếu tìm thấy
    - show: True thì lưu debug ảnh vùng tìm kiếm
    - retries: số lần thử lại
    - delay: thời gian nghỉ giữa các lần thử
    """
    for attempt in range(1, retries + 1):
        time.sleep(delay)

        matches = list(pyautogui.locateAllOnScreen(template_path, region=region, confidence=threshold))

        if matches:
            box = matches[0]   # lấy kết quả đầu tiên
            center = pyautogui.center(box)

            if click:
                pyautogui.click(center)

            # confidence (nếu có)
            conf = getattr(box, "confidence", None)
            if conf is not None:
                print(f"✅ [{attempt}/{retries}] Icon {template_path} tìm thấy tại {box}, center={center}, confidence={conf:.2f}")
            else:
                print(f"✅ [{attempt}/{retries}] Icon {template_path} tìm thấy tại {box}, center={center}")

            if show:
                screenshot = pyautogui.screenshot(region=region)
                screenshot = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
                cv2.rectangle(
                    screenshot,
                    (box.left - (region[0] if region else 0), box.top - (region[1] if region else 0)),
                    (box.left + box.width - (region[0] if region else 0), box.top + box.height - (region[1] if region else 0)),
                    (0, 0, 255),
                    2
                )
                cv2.imwrite("image/debug_pyautogui_region.png", screenshot)

            return True
        else:
            print(f"⚠️ [{attempt}/{retries}] Chưa tìm thấy icon (confidence < {threshold})")

    raise Exception(f"❌ Icon KHÔNG hiển thị sau {retries} lần thử (confidence < {threshold})")


def debug_region(image_path, region):
    x, y, w, h = region
    img = cv2.imread(image_path)

    if img is None:
        raise Exception("❌ Không tìm thấy ảnh đầu vào!")

    # Cắt vùng
    crop = img[y:y+h, x:x+w]
    cv2.imwrite("image/debug_region_crop.png", crop)

    # Vẽ khung đỏ trên ảnh gốc
    img_box = img.copy()
    cv2.rectangle(img_box, (x, y), (x+w, y+h), (0, 0, 255), 2)
    cv2.imwrite("image/debug_region_box.png", img_box)
    print()
    print("✅ Đã lưu debug_region_crop.png (ảnh cắt) và debug_region_box.png (ảnh gốc có khung đỏ)")





def get_text_default(region, lang='vie'):
    """
    Chụp ảnh màn hình, cắt vùng theo tọa độ, OCR và in ra text nhận dạng được.

    Args:
        driver: Appium webdriver đang chạy.
        region (tuple): (x1, y1, x2, y2) là vùng ảnh cần OCR.
        lang (str): Ngôn ngữ OCR ('vie' = tiếng Việt, 'eng' = tiếng Anh,...).

    Returns:
        str: Text OCR được trong vùng.
    """
    # 1️⃣ Chụp màn hình
    screenshot_path = var_stx_app.screen + "screen.png"
    var_stx_app.driver_driver.save_screenshot(screenshot_path)

    # 2️⃣ Crop vùng ảnh cần OCR
    img = Image.open(screenshot_path)
    cropped_img = img.crop(region)
    cropped_img.save(var_stx_app.screen + "cropped_region.png")  # để kiểm tra vùng cắt nếu cần

    # 3️⃣ OCR
    detected_text = pytesseract.image_to_string(cropped_img, lang=lang).strip()

    # 4️⃣ In ra text
    print(f"📌 Text OCR được trong vùng {region}: \"{detected_text}\"")
    writeData1(var_stx_app.path_luutamthoi, "Sheet1", 120, 2, detected_text)

    return detected_text



def get_text(region, lang='vie'):
    """
    Chụp ảnh màn hình, cắt vùng theo tọa độ, OCR và in ra text nhận dạng được.
    Cải tiến: tăng tương phản, chuyển grayscale, lọc màu đỏ nếu cần.
    """
    try:
        # 1️⃣ Chụp màn hình
        screenshot_path = var_stx_app.screen + "screen.png"
        var_stx_app.driver_driver.save_screenshot(screenshot_path)

        # 2️⃣ Crop vùng ảnh cần OCR
        img = Image.open(screenshot_path)
        cropped_img = img.crop(region)
        cropped_path = var_stx_app.screen + "cropped_region.png"
        cropped_img.save(cropped_path)

        # 3️⃣ Chuyển sang grayscale (xám)
        gray = cropped_img.convert("L")

        # 4️⃣ Tăng tương phản lên gấp 2.5 lần
        enhancer = ImageEnhance.Contrast(gray)
        enhanced = enhancer.enhance(2.5)

        # 5️⃣ Áp dụng threshold để tách chữ rõ hơn
        bw = enhanced.point(lambda x: 0 if x < 160 else 255, '1')

        # 6️⃣ Dự phòng: nếu chữ là màu đỏ → lọc riêng bằng OpenCV (tăng độ chính xác)
        np_img = np.array(cropped_img)
        hsv = cv2.cvtColor(np_img, cv2.COLOR_RGB2HSV)
        lower_red1 = np.array([0, 100, 100])
        upper_red1 = np.array([10, 255, 255])
        lower_red2 = np.array([160, 100, 100])
        upper_red2 = np.array([179, 255, 255])
        mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
        mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
        mask = cv2.bitwise_or(mask1, mask2)
        red_only = cv2.bitwise_and(np_img, np_img, mask=mask)
        red_path = var_stx_app.screen + "red_only.png"
        cv2.imwrite(red_path, red_only)

        # 7️⃣ OCR với cấu hình chính xác cao
        custom_config = r'--oem 3 --psm 7'
        detected_text = pytesseract.image_to_string(bw, lang=lang, config=custom_config).strip()

        # Nếu OCR trắng đen rỗng → thử OCR trên ảnh lọc đỏ
        if not detected_text:
            detected_text = pytesseract.image_to_string(red_only, config=custom_config).strip()

        # 8️⃣ In ra kết quả
        print(f"📌 Text OCR được trong vùng {region}: \"{detected_text}\"")

        # 9️⃣ Ghi ra file Excel
        writeData1(var_stx_app.path_luutamthoi, "Sheet1", 120, 2, detected_text)

        return detected_text

    except Exception as e:
        print(f"❌ Lỗi OCR: {e}")
        return ""








import subprocess
import re
import time

# ✅ Đường dẫn adb LDPlayer (cố định)
ADB_PATH = r"D:\LDPlayer\LDPlayer9\adb.exe"

def get_first_device():
    """Lấy ID thiết bị ADB đầu tiên (vd: emulator-5556)."""
    output = subprocess.check_output([ADB_PATH, "devices"], text=True)
    lines = output.strip().splitlines()[1:]  # Bỏ dòng đầu "List of devices attached"
    for line in lines:
        if line.strip() and "device" in line:
            return line.split()[0]
    return None

def adb_shell(cmd, device_id=None):
    """Chạy lệnh ADB shell trên thiết bị chỉ định."""
    if not device_id:
        device_id = get_first_device()
    if not device_id:
        raise RuntimeError("❌ Không tìm thấy thiết bị ADB nào đang kết nối.")
    full_cmd = [ADB_PATH, "-s", device_id, "shell"] + cmd.split()
    print(f"👉 Running: {' '.join(full_cmd)}")
    return subprocess.check_output(full_cmd, text=True).strip()

def check_devices():
    """Liệt kê các thiết bị ADB đang kết nối."""
    output = subprocess.check_output([ADB_PATH, "devices"], text=True)
    print(output)




def swipe_percent(driver, start_x_pct, start_y_pct, end_x_pct, end_y_pct, duration=500):
    size = driver.get_window_size()
    width = size['width']
    height = size['height']

    start_x = int(width * start_x_pct)
    start_y = int(height * start_y_pct)
    end_x = int(width * end_x_pct)
    end_y = int(height * end_y_pct)

    driver.swipe(start_x, start_y, end_x, end_y, duration)


def tap_percent(driver, x_pct, y_pct):
    size = driver.get_window_size()
    width = size['width']
    height = size['height']

    x = int(width * x_pct)
    y = int(height * y_pct)

    driver.tap([(x, y)])






def close_chrome():
    print("đang đóng trang")
    web = homepage_g7.web_app()
    web.create_driver()
    web.driver2.quit()










@retry(tries=3, delay=2, backoff=1, jitter=5, )
def close_app(driver):
    try:
        driver.implicitly_wait(0.5)
    except:
        var_stx_app.restart_driver(driver)#kết nối lại adb
        driver.implicitly_wait(0.5)

    homepage_g7.homepage_g7_back()
    homepage_driver.homepage_g7_drive_back()
    # clear_all()



























def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_stx_app.timerun)
        writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        if var_stx_app.timerun == "":
            writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        else:
            writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, var_stx_app.timerun)


        if var_stx_app.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_stx_app.timerun == "":
            break





def clearData(file,sheetName, data, ketqua, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = data
        sheet["G"+i] = ketqua
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)


def clearData_luutamthoi(file,sheetName, overview, detail, api):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 1
    while (i < 300):
        i += 1
        i = str(i)
        sheet["B"+i] = overview
        sheet["C"+i] = detail
        sheet["D"+i] = api
        i = int(i)
    wordbook.save(file)




from filelock import FileLock
import openpyxl
import time
import uuid
import shutil
import os

RETRY_DELAY = 0.5

def _safe_save(workbook, file):
    tmp = f"{file}.{uuid.uuid4().hex}.tmp"
    workbook.save(tmp)
    shutil.move(tmp, file)




def readData(file, sheetName, rownum, columnno):
    lock = FileLock(file + ".lock")
    while True:
        try:
            with lock:
                wb = openpyxl.load_workbook(file)
                try:
                    sheet = wb[sheetName]
                    return sheet.cell(row=rownum, column=columnno).value
                finally:
                    wb.close()
        except Exception as e:
            if "zip file" in str(e).lower():
                time.sleep(RETRY_DELAY)
            else:
                raise




def writeData(file, sheetName, caseid, columnno, data):
    lock = FileLock(file + ".lock")
    while True:
        try:
            with lock:
                wb = openpyxl.load_workbook(file)
                try:
                    sheet = wb[sheetName]

                    for i in range(1, 5001):
                        if sheet[f"A{i}"].value == caseid:
                            sheet.cell(row=i, column=columnno).value = data
                            break

                    _safe_save(wb, file)
                    return
                finally:
                    wb.close()
        except Exception as e:
            if "zip file" in str(e).lower():
                time.sleep(RETRY_DELAY)
            else:
                raise




def writeData1(file, sheetName, rownum, columnno, data):
    lock = FileLock(file + ".lock")
    while True:
        try:
            with lock:
                wb = openpyxl.load_workbook(file)
                try:
                    sheet = wb[sheetName]
                    sheet.cell(row=rownum, column=columnno).value = data

                    _safe_save(wb, file)
                    return
                finally:
                    wb.close()
        except Exception as e:
            if "zip file" in str(e).lower():
                time.sleep(RETRY_DELAY)
            else:
                raise






# def readData(file,sheetName,rownum,columnno):
#     with excel_lock:
#         wordbook = openpyxl.load_workbook(file)
#         sheet = wordbook.get_sheet_by_name(sheetName)
#         return sheet.cell(row=rownum, column=columnno).value


# def writeData(file,sheetName,caseid,columnno,data):
#     wordbook = openpyxl.load_workbook(file)
#     sheet = wordbook.get_sheet_by_name(sheetName)
#     i = 0
#     while (i < 5000):
#         i += 1
#         i = str(i)
#         if sheet["A"+i].value == caseid:
#             i = int(i)
#             sheet.cell(row=i, column=columnno).value = data
#             break
#         i = int(i)
#     wordbook.save(file)



# def writeData1(file,sheetName,rowum,columnno,data):
#     with excel_lock:
#         wordbook = openpyxl.load_workbook(file)
#         sheet = wordbook.get_sheet_by_name(sheetName)
#         sheet.cell(row=rowum, column=columnno).value = data
#         wordbook.save(file)






def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_stx_app.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)
    module_stx_app.check_casenone()
    module_stx_app.check_casefail()
    module_stx_app.check_casepass()
    module_stx_app.check_casespecial()


    luong = var_stx_app.modetest
    if luong == "0":
        luong = "Luồng Full(0)"

    if luong == "1":
        luong = "Luồng khách hàng(1)"

    if luong == "2":
        luong = "Luồng tài xế(2)"

    mucnghiemtrong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 85, 2))
    tong_case_trong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 86, 2))
    case_fail = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 167, 2))
    case_pass = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 177, 2))

    print("Pass: "+ case_pass)
    print("Fail: "+ case_fail)
    print("Trống: "+ tong_case_trong)
    print("Nghiêm trọng: "+ mucnghiemtrong)




    thoigianbatdauchay = str(readData(var_stx_app.path_luutamthoi, "Sheet2", 2, 2))

    print(thoigianbatdauchay)
    # if case_fail >= 1:
    driver2.get("https://web.telegram.org/a/#-4248738484")
    time.sleep(2)
    case_pass = str(case_pass)
    case_fail = str(case_fail)
    driver2.ele(var_stx_app.hopthoai).click()
    time.sleep(0.5)
    time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    driver2.ele(var_stx_app.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- ModeTest: " + luong+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)
    driver2.ele(var_stx_app.hopthoai_input).input(Keys.ENTER)
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink_file).click()
    time.sleep(1)
    subprocess.Popen(var_stx_app.uploadpath+"checklist_app_staxi.exe")
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_send).click()
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink_file).click()
    time.sleep(2)
    subprocess.Popen(var_stx_app.uploadpath+"app_staxi_log.exe")
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_send).click()
    time.sleep(1)





def viber_send_text():
    module_stx_app.check_casenone()
    module_stx_app.check_casefail()
    module_stx_app.check_casepass()
    module_stx_app.check_casespecial()

    luong = var_stx_app.modetest
    if luong == "0":
        luong = "Luồng Full(0)"

    if luong == "1":
        luong = "Luồng tương tác 2app(1)"

    if luong == "2":
        luong = "Luồng khách hàng(2)"

    if luong == "3":
        luong = "Luồng tài xế(3)"

    mucnghiemtrong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 85, 2))
    tong_case_trong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 86, 2))
    case_fail = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 167, 2))
    case_pass = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 177, 2))

    version_g7 = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 122, 2))
    version_driver = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 123, 2))



    print("Pass: "+ case_pass)
    print("Fail: "+ case_fail)
    print("Trống: "+ tong_case_trong)
    print("Nghiêm trọng: "+ mucnghiemtrong)
    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)

    thoigianbatdauchay = str(readData(var_stx_app.path_luutamthoi, "Sheet2", 2, 2))


    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- ModeTest: " + luong+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)

    AUTH_TOKEN = "54c429aa33f11085-5f24879e97bb7755-8f4f9f372469b338"  # id Cảnh báo Autotest App STAXI Customer
    FROM_USER_ID = "ft4+qte2zSLstobjJFTt8w=="


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    print(AUTH_TOKEN)
    print(FROM_USER_ID)


    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return



    # 2. Gửi tin nhắn văn bản
    send_url = "https://chatapi.viber.com/pa/post"
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,
        "type": "text",
        "text": ("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- Version G7 Taxi: " + version_g7 +
                                              "\n- Version Lái xe G7: " + version_driver +
                                              "\n- ModeTest: " + luong+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)}

    headers = {
        "Content-Type": "application/json"}

    response = requests.post(send_url, json=payload, headers=headers)

    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())




def check_user_id():
    res = requests.post(
        "https://chatapi.viber.com/pa/get_account_info",
        headers={"X-Viber-Auth-Token": "54c429aa33f11085-5f24879e97bb7755-8f4f9f372469b338"}
    )
    print(res.json())





def upload_pixeldrain_auth(filepath):
    API_KEY = "c567bb13-f4c0-4aac-b9bd-c8add1e467fc"  # Thay bằng key thật

    with open(filepath, "rb") as f:
        res = requests.post(
            "https://pixeldrain.com/api/file",
            files={"file": f},
            auth=HTTPBasicAuth('', API_KEY)
        )
        res_json = json.loads(res.text)
        file_id = res_json["id"]
        link_download = (f"https://pixeldrain.com/api/file/{file_id}")
        print(link_download)
        return link_download




def send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, file_path):
    file_url = upload_pixeldrain_auth(file_path)
    if not file_url:
        print("⚠️ Không thể upload file. Hủy gửi.")
        return

    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,           # Viber user ID người nhận
        "min_api_version": 1,
        "tracking_data": "",               # Có thể để chuỗi rỗng nếu không dùng tracking
        "type": "file",
        "media": file_url,
        "size": file_size,
        "file_name": file_name}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())



def viber_send_file():
    # ==== Ví dụ sử dụng ====
    AUTH_TOKEN = "54c429aa33f11085-5f24879e97bb7755-8f4f9f372469b338"  # id Cảnh báo Autotest App STAXI
    FROM_USER_ID = "ft4+qte2zSLstobjJFTt8w=="


    FILE_PATH_checklisst = var_stx_app.checklistpath  # Thay bằng đường dẫn file thật
    FILE_PATH_log = var_stx_app.logpath  # Thay bằng đường dẫn file thật

    print(var_stx_app.checklistpath)
    print(var_stx_app.logpath)


    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_checklisst)
    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_log)



def send_viber():
    viber_send_text()
    viber_send_file()




def delete_image():
    path = os.path.join(var_stx_app.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def get_datachecklist(ma):
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 3000):
        rownum += 1
        rownum = str(rownum)
        if sheet["A" + rownum].value == ma:
            tensukien = sheet["B" + rownum].value
            ketqua = sheet["E" + rownum].value
            print(tensukien)
            print(ketqua)
            writeData1(var_stx_app.path_luutamthoi, "Sheet1", 42, 2, tensukien)
            writeData1(var_stx_app.path_luutamthoi, "Sheet1", 43, 2, ketqua)
        rownum = int(rownum)



def write_caseif():
    n = 0
    while (n < 11):
        n += 1
        n = str(n)
        print("try:\n   if case == 'Setting"+n+"':\n       caseid_stx_app.caseid_setting"+n+"(self)\nexcept:\n    module_other_stx_app.close_app()")
        n = int(n)




def write_caseif1():
    n = 1
    while (n < 36):
        n += 1
        n = str(n)
        print("try:\n   caseid_app.caseid_utilities"+n+"(self)\nexcept:\n     module_other_app.close_app()")
        n = int(n)
        



def write_caseif2():
    n = 84
    while (n < 155):
        n += 1
        n = str(n)
        print("caseid_app.caseid_minitor"+n+"(self='""')")
        n = int(n)
#caseid_app.caseid_minitor83(self="")





def write_caseif3():
    n = 0
    while (n < 17):
        n += 1
        n = str(n)
        print("try:\n   if case == 'Report158_"+n+"':\n       login_app.login.logout_back(self)\n       caseid_app.caseid_report158_"+n+"(self)  \nexcept:\n    module_other_app.close_app()")
        n = int(n)





def write_result_text_try_if_cut(driver, code, eventname, result, path_module, path_text, number_from, number_to, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        var_stx_app.logging.info(check_text[number_from: number_to])
        var_stx_app.logging.info(check_result)
        if check_text[number_from: number_to] == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_cut(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, 10, 15, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")


def write_result_text_try_if1(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        check_text = action.get_text(driver, path_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_display_image(driver, code, eventname, result, path_module, path_icon, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)

    try:
        check_icon(var_stx_app.screen + path_icon, region=None, threshold=0.8, click=False, show=True)
        # homepage_driver.check_image(3, var_stx_app.screen+path_icon, 0.8, False)
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_display_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_not_display_image(driver, code, eventname, result, path_module, path_icon, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)

    try:
        check_icon(var_stx_app.screen + path_icon, region=None, threshold=0.8, click=False, show=True)
        # homepage_driver.check_image(3, var_stx_app.screen+path_icon, 0.8, False)
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")






def write_result_text_try_if_image_other(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        region = (x1, y1, x2, y2)
        get_text(region)
        check_text = get_text_default(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")


def write_result_text_try_if_image_other1(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        region = (x1, y1, x2, y2)
        get_text(region)
        check_text = get_text(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_displayed_try_image(driver, code, eventname, result, path_module, image, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        homepage_driver.check_image(1, f"{image}.png", 0.8, True)
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_displayed_try_image(code, eventname, result, "Quản trị - Danh sách xe",
    #                                "tenanh", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_image(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        region = (x1, y1, x2, y2)
        get_text(region)
        check_text = get_text_default(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_image_cv(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        region = (x1, y1, x2, y2)
        get_text(region)
        check_text = get_text(region)
        print("👉 Text nhận dạng được:", check_text)
        check_text = homepage_driver.normalize_text(check_text)

        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_image_not_in(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    try:
        region = (x1, y1, x2, y2)
        get_text(region)
        check_text = get_text(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result not in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)





def write_result_text_try_if_other_image(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    region = (x1, y1, x2, y2)
    get_text(region)

    try:
        check_text = get_text_default(region)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_other_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_image_in(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    region = (x1, y1, x2, y2)
    get_text(region)

    try:
        check_text = get_text_default(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_image_in1(driver, code, eventname, result, path_module, x1, y1, x2, y2, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    # region = (168, 1319, 550, 1355)
    region = (x1, y1, x2, y2)
    get_text(region)

    try:
        check_text = get_text_default(region)
        print("👉 Text nhận dạng được:", check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text in check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_image(code, eventname, result, "Quản trị - Danh sách xe", 168, 1319, 550, 1355,
    #                                "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_in(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)


def write_result_text_try_if_in1(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text in check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if_toast(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, "//android.widget.Toast").text
        print(check_text)
        # check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_toast_in(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, "//android.widget.Toast").text
        print(check_text)
        # check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_toast_other(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, "//android.widget.Toast").text
        print(check_text)
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_toast_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")






def write_result_text_try_if_not_fail(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_not_fail_other(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_conten(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_other(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_other_not_in(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_result not in check_text:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_conten_other(driver, code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")








def write_result_displayed_try(driver, code, eventname, result, path_module, path_text, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        driver.find_element(By.XPATH, path_text).is_displayed()
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")



def write_result_not_displayed_try(driver, code, eventname, result, path_module, path_text, name_image):
    driver.implicitly_wait(2)
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_not_displayed = driver.find_element(By.XPATH, path_text).is_displayed()
        var_stx_app.logging.info("False")
        driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")

        
        
        
def write_result_excelreport(driver, code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_stx_app.logging.info("Check vị trí "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")



def write_result_excelreport_clear_data(driver, code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_stx_app.logging.info("Check vị trí "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, " ")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")

        


