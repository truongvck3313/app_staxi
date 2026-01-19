import time
import openpyxl
import os
from selenium.common.exceptions import NoSuchElementException
import subprocess
import var_stx_app
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import logging
from retry import retry
import module_stx_app
import homepage_g7
import homepage_driver
from pynput.keyboard import Key, Controller
import pygetwindow as gw
import pyautogui
from pynput.keyboard import Controller, Key
import time
import re


keyboard = Controller()

def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename="bagps_app.log",
                                                     encoding='utf-8', mode='w')],  #w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)





def check_devices():
    adb_path = r"D:\LDPlayer\LDPlayer9\adb.exe"

    # Liệt kê các thiết bị kết nối
    output = subprocess.check_output([adb_path, "devices"], text=True)
    print(output)



def reset_driver():

    # Đường dẫn adb của LDPlayer
    # adb_path = r"D:\LDPlayer\LDPlayer9\adb.exe"

    adb_path = r"{}adb.exe".format(var_stx_app.ldp)
    emulator_name = "emulator-5556"  # LDPlayer mặc định
    print(adb_path)

    # Hàm chạy lệnh ADB
    def adb_shell(cmd):
        return subprocess.check_output([adb_path, "-s", emulator_name, "shell"] + cmd.split(), text=True)

    # Lấy danh sách package đang chạy (foreground hoặc recent)
    print("Lấy danh sách app đang chạy...")
    output = adb_shell("dumpsys activity recents | grep 'Recent #'")

    # Lọc ra các gói package (dòng có dạng: Recent #0: TaskRecord{... A=com.example.app U=0 StackId=1})
    package_names = re.findall(r"A=([\w\.]+)", output)
    package_names = list(set(package_names))  # Loại bỏ trùng

    print(f"Đã tìm thấy {len(package_names)} app đang chạy:")
    for pkg in package_names:
        print("  -", pkg)

    # Force-stop từng app
    for pkg in package_names:
        print(f"➤ Dừng app: {pkg}")
        adb_shell(f"am force-stop {pkg}")
        time.sleep(0.5)

    # Tùy chọn: khởi động lại app chính (nếu biết package/activity)
    # Ví dụ với BA GPS
    main_apps = {
        "com.bagps": "com.bagps.MainActivity",  # bạn có thể thêm các app khác nếu cần
    }

    for pkg, activity in main_apps.items():
        print(f"🚀 Khởi động lại app: {pkg}")
        adb_shell(f"am start -n {pkg}/{activity}")
        time.sleep(1)

    print("✅ Reset tất cả app đã xong.")






@retry(tries=3, delay=2, backoff=1, jitter=5, )
def close_app():
    var_stx_app.driver.implicitly_wait(0.5)
    homepage_g7.homepage_g7_back()
    homepage_driver.homepage_g7_drive_back()
    clear_all()




def check_error():
    ld = gw.getWindowsWithTitle('LDPlayer')
    print("check_eror0.0")
    if ld:
        print("check_eror0.1")
        ld[0].activate()  # đưa lên foreground
        print("check_eror0.2")
        time.sleep(1)  # đợi cho nó ổn định
        print("đã tìm thấy LDPlayer")
    else:
        print("Không tìm thấy LDPlayer")

    print("check_eror01")
    keyboard.press(Key.f2)
    time.sleep(2)
    print("check_eror02")
    var_stx_app.driver.swipe(420, 200, 420, 1350, 300)
    time.sleep(2)
    print("check_eror03")
    var_stx_app.driver.find_element(By.XPATH, var_stx_app.CLEAR_ALL).click()
    time.sleep(2)
    print("check_eror04")






def clear_all():
    window = gw.getWindowsWithTitle('LDPlayer')[0]
    window.activate()
    time.sleep(1)
    try:
        var_stx_app.driver.implicitly_wait(5)
        var_stx_app.driver.find_element(By.XPATH, var_stx_app.close_app).click()
        time.sleep(3)
    except:
        pass


    n = 0
    while (n < 4):
        n += 1
        print("đóng app lần: {}".format(n))
        try:
            print("đóng app c1 lần thứ: {}".format(n))
            print("clear_all_1")
            keyboard.press(Key.f2)
            print("clear_all_2")
            time.sleep(2)
            var_stx_app.driver.swipe(420, 200, 420, 1350, 300)
            print("clear_all_3")
            time.sleep(2)
            var_stx_app.driver.find_element(By.XPATH, var_stx_app.CLEAR_ALL).click()
            print("clear_all_4")
            time.sleep(2)
            print("đóng app c1")
            break
        except:
            pass
        try:
            print("đóng app c2 lần thứ: {}".format(n))
            print("clear_all_5")
            pyautogui.press('f2')
            print("clear_all_6")
            time.sleep(2)
            var_stx_app.driver.swipe(420, 200, 420, 1350, 300)
            print("clear_all_7")
            time.sleep(2)
            var_stx_app.driver.find_element(By.XPATH, var_stx_app.CLEAR_ALL).click()
            time.sleep(2)
            print("đóng app c2")
            break
        except:
            pass

        if n == 3:
            print("clear_all_9")
            reset_driver()
            print("clear_all_10")
            break
        if n == 4:
            print("clear_all_11")
            close_app_image()




def click_app(name_app):
    n = 0
    while (n < 25):
        n += 1
        try:
            keyboard.press(Key.f1)
            time.sleep(1)
            keyboard.press(Key.f1)
            time.sleep(2)
            if name_app == "G7 Taxi":
                var_stx_app.driver.find_element(By.XPATH, var_stx_app.g7_taxi).click()
                time.sleep(3)
                print("đã vào app khách")
                break
            if name_app == "Lái xe G7":
                var_stx_app.driver.find_element(By.XPATH, var_stx_app.driver_g7).click()
                time.sleep(3)
                print("đã vào app xế")
                break
            print("đang ở vòng: {}".format(n))
        except:
            pass




def setup_app_g7():
    try:
        var_stx_app.driver.implicitly_wait(0.3)
        var_stx_app.driver.find_element(By.XPATH, var_stx_app.g7_taxi).click()
        time.sleep(5)
    except:
        pass

    try:
        var_stx_app.driver.implicitly_wait(0.3)
        var_stx_app.driver.find_element(By.XPATH, var_stx_app.homepage_iconx1).click()
        time.sleep(1.5)
    except:
        pass





def close_app_image():
    import pyautogui
    import time
    # Click vào nút "Khởi động lại" 1
    try:
        print("close_app_image_01")
        clear_button1 = pyautogui.locateOnScreen(var_stx_app.uploadpath+'clear_button1.png', confidence=0.8)
        print("close_app_image_02")
        if clear_button1:
            print("close_app_image_03")
            pyautogui.click(clear_button1)
            print("close_app_image_04")
            time.sleep(2)
            var_stx_app.driver.swipe(420, 200, 420, 1350, 300)
            time.sleep(2)
            print("close_app_image_05")
            var_stx_app.driver.find_element(By.XPATH, var_stx_app.CLEAR_ALL).click()
            print("close_app_image_06")
            time.sleep(2)
        else:
            print("Không tìm thấy nút đóng app! 1")
    except:
        print("đang ẩn app hoặc chưa đóng pycharm nên không đóng được app 1")


    # # Click vào nút "Khởi động lại" 2
    # try:
    #     clear_button2 = pyautogui.locateOnScreen(var_stx_app.uploadpath+'clear_button1.png', confidence=0.8)
    #     if clear_button2:
    #         pyautogui.click(clear_button2)
    #         time.sleep(2)
    #         var_stx_app.driver.swipe(420, 200, 420, 1350, 300)
    #         time.sleep(2)
    #         var_stx_app.driver.find_element(By.XPATH, var_stx_app.CLEAR_ALL).click()
    #         time.sleep(2)
    #     else:
    #         print("Không tìm thấy nút đóng app! 2")
    # except:
    #     print("đang ẩn app hoặc chưa đóng pycharm nên không đóng được app 2")










def fake_ip():
    from faker import Faker
    from faker.providers import internet
    fake = Faker()
    fake.add_provider(internet)
    print(fake.ipv4_private())





def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_stx_app.timerun)
        writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        if var_stx_app.timerun == "":
            writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        else:
            writeData1(var_stx_app.path_luutamthoi, "Sheet2", 2, 2, var_stx_app.timerun)


        if var_stx_app.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_stx_app.timerun == "":
            break





def clearData(file,sheetName, data, ketqua, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = data
        sheet["G"+i] = ketqua
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)


def clearData_luutamthoi(file,sheetName, overview, detail, api):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 1
    while (i < 300):
        i += 1
        i = str(i)
        sheet["B"+i] = overview
        sheet["C"+i] = detail
        sheet["D"+i] = api
        i = int(i)
    wordbook.save(file)






def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value




def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)




def writeData1(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)





def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_stx_app.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)
    module_stx_app.check_casenone()
    module_stx_app.check_casefail()
    module_stx_app.check_casepass()


    luong = var_stx_app.modetest
    if luong == "0":
        luong = "Luồng Full(0)"

    if luong == "1":
        luong = "Luồng khách hàng(1)"

    if luong == "2":
        luong = "Luồng tài xế(2)"

    mucnghiemtrong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 85, 2))
    tong_case_trong = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 86, 2))
    case_fail = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 167, 2))
    case_pass = str(readData(var_stx_app.path_luutamthoi, 'Sheet1', 177, 2))

    print("Pass: "+ case_pass)
    print("Fail: "+ case_fail)
    print("Trống: "+ tong_case_trong)
    print("Nghiêm trọng: "+ mucnghiemtrong)




    thoigianbatdauchay = str(readData(var_stx_app.path_luutamthoi, "Sheet2", 2, 2))

    print(thoigianbatdauchay)
    # if case_fail >= 1:
    driver2.get("https://web.telegram.org/a/#-4248738484")
    time.sleep(2)
    case_pass = str(case_pass)
    case_fail = str(case_fail)
    driver2.ele(var_stx_app.hopthoai).click()
    time.sleep(0.5)
    time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    driver2.ele(var_stx_app.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- ModeTest: " + luong+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)
    driver2.ele(var_stx_app.hopthoai_input).input(Keys.ENTER)
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink_file).click()
    time.sleep(1)
    subprocess.Popen(var_stx_app.uploadpath+"checklist_app_staxi.exe")
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_send).click()
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_stx_app.hopthoai_iconlink_file).click()
    time.sleep(2)
    subprocess.Popen(var_stx_app.uploadpath+"app_staxi_log.exe")
    time.sleep(2)
    driver2.ele(var_stx_app.hopthoai_send).click()
    time.sleep(1)




def delete_image():
    path = os.path.join(var_stx_app.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def get_datachecklist(ma):
    wordbook = openpyxl.load_workbook(var_stx_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 3000):
        rownum += 1
        rownum = str(rownum)
        if sheet["A" + rownum].value == ma:
            tensukien = sheet["B" + rownum].value
            ketqua = sheet["E" + rownum].value
            print(tensukien)
            print(ketqua)
            writeData1(var_stx_app.path_luutamthoi, "Sheet1", 42, 2, tensukien)
            writeData1(var_stx_app.path_luutamthoi, "Sheet1", 43, 2, ketqua)
        rownum = int(rownum)



def write_caseif():
    n = 0
    while (n < 11):
        var_stx_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Setting"+n+"':\n       caseid_stx_app.caseid_setting"+n+"(self)\nexcept:\n    module_other_stx_app.close_app()")
        n = int(n)




def write_caseif1():
    n = 1
    while (n < 36):
        var_stx_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   caseid_app.caseid_utilities"+n+"(self)\nexcept:\n     module_other_app.close_app()")
        n = int(n)
        



def write_caseif2():
    n = 84
    while (n < 155):
        var_stx_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("caseid_app.caseid_minitor"+n+"(self='""')")
        n = int(n)
#caseid_app.caseid_minitor83(self="")





def write_caseif3():
    n = 0
    while (n < 17):
        var_stx_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Report158_"+n+"':\n       login_app.login.logout_back(self)\n       caseid_app.caseid_report158_"+n+"(self)  \nexcept:\n    module_other_app.close_app()")
        n = int(n)





def write_result_text_try_if_cut(code, eventname, result, path_module, path_text, number_from, number_to, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        var_stx_app.logging.info(check_text[number_from: number_to])
        var_stx_app.logging.info(check_result)
        if check_text[number_from: number_to] == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_cut(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, 10, 15, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_toast(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, "//android.widget.Toast").text
        print(check_text)
        # check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_not_fail(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_not_fail_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_conten(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text == check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_conten_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_stx_app.driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_stx_app.checklistpath, "Checklist", code, 6, check_text)
        var_stx_app.logging.info(check_text)
        if check_text != check_result:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")








def write_result_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        var_stx_app.driver.find_element(By.XPATH, path_text).is_displayed()
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")



def write_result_not_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_stx_app.driver.implicitly_wait(2)
    var_stx_app.logging.info("--------------")
    var_stx_app.logging.info(path_module)
    var_stx_app.logging.info("Mã - " + code)
    var_stx_app.logging.info("Tên sự kiện - " + eventname)
    var_stx_app.logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var_stx_app.driver.find_element(By.XPATH, path_text).is_displayed()
        var_stx_app.logging.info("False")
        var_stx_app.driver.save_screenshot(var_stx_app.imagepath + code + name_image)
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx_app.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        var_stx_app.logging.info("True")
        writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")

        
        
        
def write_result_excelreport(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_stx_app.logging.info("Check vị trí "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")



def write_result_excelreport_clear_data(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_stx_app.logging.info("Check vị trí "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_stx_app.logging.info("True")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, " ")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_stx_app.logging.info("False")
            writeData(var_stx_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")

        


